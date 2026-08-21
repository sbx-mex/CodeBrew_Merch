"""Pruebas de contrato para motores Excel y catálogo visual."""

from __future__ import annotations

import json
import csv
import hashlib
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from scripts.integrate_uploaded_images import add_woe_merch_products, discover_images, identifier, validate_published_lots
from scripts.generate_products import workbook_modified_at


ROOT = Path(__file__).resolve().parents[1]


def load_catalog() -> dict:
    text = (ROOT / "data/merch-catalog.js").read_text(encoding="utf-8")
    return json.loads(text.split("=", 1)[1].strip().rstrip(";"))


class VisualCatalogContractTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.catalog = load_catalog()
        cls.products = cls.catalog["products"]

    def test_three_independent_engines_are_reported(self) -> None:
        self.assertEqual(self.catalog["meta"]["engineFiles"], 3)
        self.assertEqual(self.catalog["meta"]["visualSourceFiles"], 3)
        self.assertEqual(len(list((ROOT / "engines/merch-lists").glob("*.xlsx"))), 3)
        self.assertEqual(len(list((ROOT / "engines/visual-sources").glob("*.zip"))), 3)

    def test_build_timestamp_comes_from_workbook_not_checkout_mtime(self) -> None:
        report = json.loads((ROOT / "data/import-report.json").read_text(encoding="utf-8"))
        self.assertEqual(report["generatedAtUtc"], workbook_modified_at(ROOT / "Lista_Precios_Base.xlsx"))

    def test_stable_article_keys_are_unique(self) -> None:
        keys = [product["articleKey"] for product in self.products]
        self.assertTrue(keys)
        self.assertEqual(len(keys), len(set(keys)))
        self.assertTrue(all(key.startswith("dia-") and "--pos-" in key for key in keys))

    def test_every_product_has_operational_fields_and_photo_state(self) -> None:
        for product in self.products:
            self.assertTrue(product["codigoDia"])
            self.assertTrue(product["displayName"])
            self.assertTrue(product["nameKey"])
            day_key = identifier(product["codigoDia"])
            self.assertEqual(product.get("photoUploadName"), f"{day_key}.jpg" if day_key else "")
            self.assertNotIn("prices", product)
            if product.get("visual"):
                self.assertEqual(product.get("visualSource"), "manual-upload")
                self.assertTrue(product["visual"]["src"].startswith("assets/catalog/images/lote-"))
            else:
                self.assertEqual(product.get("visualSource"), "pending-upload")
            self.assertIn(product.get("stockPriority"), {"active", "secondary"})

    def test_display_names_keep_sap_and_micros_source_priority(self) -> None:
        crossed = [product for product in self.products if product.get("sapDescriptions") or product.get("microsNames")]
        self.assertTrue(crossed)
        app = (ROOT / "app.js").read_text(encoding="utf-8")
        self.assertIn("inventory=p.microsNames?.[0]", app)
        self.assertIn("sci=p.sapDescriptions?.[0]", app)

    def test_github_limits_are_respected(self) -> None:
        excluded = {".git", ".codebrew-build", "__pycache__"}
        for path in ROOT.rglob("*"):
            relative = path.relative_to(ROOT)
            if any(part in excluded for part in relative.parts):
                continue
            if path.is_file():
                self.assertLess(path.stat().st_size, 25_000_000, relative.as_posix())
            elif path.is_dir():
                self.assertLessEqual(sum(child.is_file() for child in path.iterdir()), 100, relative.as_posix())

    def test_four_manual_image_lots_are_integrated(self) -> None:
        lots = sorted(path for path in (ROOT / "assets/catalog/images").iterdir() if path.is_dir())
        restored = [path for lot in lots for path in lot.iterdir() if path.is_file() and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}]
        featured = list((ROOT / "assets/catalog/featured").glob("*.webp")) if (ROOT / "assets/catalog/featured").exists() else []
        self.assertEqual([lot.name for lot in lots], ["lote-01", "lote-02", "lote-03", "lote-04"])
        counts = [sum(path.is_file() and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"} for path in lot.iterdir()) for lot in lots]
        self.assertTrue(all(count <= 100 for count in counts))
        # Los cuatro lotes son fuentes independientes; no se exige llenar el
        # anterior antes de publicar el siguiente.
        self.assertEqual(len(restored), len({path.stem for path in restored}))
        self.assertEqual(featured, [])
        self.assertEqual(self.catalog["meta"].get("imageMode"), "manual-upload")
        self.assertEqual(self.catalog["meta"].get("publishedImageFiles"), len(restored))
        coverage = json.loads((ROOT / "data/photo-coverage.json").read_text(encoding="utf-8"))
        self.assertEqual(self.catalog["meta"].get("unmatchedImageFiles"), coverage["totals"]["pendingRelationImageFiles"])
        self.assertEqual(coverage["totals"]["publishedImageFiles"], coverage["totals"]["matchedImageFiles"] + coverage["totals"]["pendingRelationImageFiles"])
        self.assertEqual(coverage["unmatchedPolicy"], "reject-and-report-do-not-guess")
        self.assertEqual(coverage["publishedNaming"], "codigo-dia.webp")
        self.assertEqual(coverage["totals"]["pendingRelationImageFiles"], 0)
        self.assertTrue(all(path.suffix == ".webp" and path.stem.isdigit() for path in restored))
        for path in restored:
            with Image.open(path) as image:
                self.assertEqual(image.size, (960, 960), path.name)
        self.assertLessEqual(sum(path.stat().st_size for path in restored), 8_000_000)
        self.assertTrue(all(row["status"] in {"matched", "pending-match", "ignored-duplicate-article", "ignored-duplicate-content"} for row in coverage["files"]))
        for codigo in ("16889", "16972", "16990", "17336", "17337", "17338"):
            self.assertTrue((ROOT / f"assets/catalog/images/lote-02/{codigo}.webp").is_file())
            self.assertFalse((ROOT / f"assets/catalog/images/lote-01/{codigo}.webp").exists())
        self.assertEqual(self.catalog["meta"].get("withSourceImage"), sum(bool(product.get("visual")) for product in self.products))

    def test_missing_empty_lots_are_created_in_a_clean_checkout(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "images"
            images, lots = discover_images(source)
            self.assertEqual(images, [])
            self.assertEqual([row["folder"] for row in lots], ["lote-01", "lote-02", "lote-03", "lote-04"])
            self.assertTrue(all((source / row["folder"]).is_dir() for row in lots))

    def test_newer_lot_supersedes_duplicate_from_earlier_lot(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "images"
            for number in range(1, 5):
                (source / f"lote-{number:02d}").mkdir(parents=True, exist_ok=True)
            image = Image.new("RGB", (20, 20), "green")
            image.save(source / "lote-01/16999.jpg")
            image.save(source / "lote-02/16999.jpg")
            images, lots = discover_images(source)
            self.assertEqual([path.relative_to(source).as_posix() for path in images], ["lote-02/16999.jpg"])
            self.assertEqual(sum(row["duplicatesIgnored"] for row in lots), 1)

    def test_equal_content_with_distinct_identifiers_is_preserved(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "images"
            for number in range(1, 5):
                (source / f"lote-{number:02d}").mkdir(parents=True, exist_ok=True)
            image = Image.new("RGB", (20, 20), "green")
            image.save(source / "lote-01/16999.jpg")
            image.save(source / "lote-03/11157547.jpg")
            images, lots = discover_images(source)
            self.assertEqual(
                [path.relative_to(source).as_posix() for path in images],
                ["lote-01/16999.jpg", "lote-03/11157547.jpg"],
            )
            self.assertEqual(lots[2]["duplicatesIgnored"], 0)

    def test_exact_sap_day_with_photo_is_added_without_micros_classification(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            woe_path = Path(temp) / "woe.js"
            woe_path.write_text(
                'window.WOE_CATALOG = [{"idWoe":"155962","codigoDia":"16960",'
                '"descripcionSap":"60979 C2P CORE PLSTC HTCDCP DCHRC 24OZ",'
                '"micros":[],"merchCategory":"","sourceRow":1573}];\nwindow.WOE_META = {};\n',
                encoding="utf-8",
            )
            products: list[dict] = []
            self.assertEqual(add_woe_merch_products(products, woe_path, {"16960"}), 1)
            self.assertEqual(products[0]["codigoDia"], "16960")
            self.assertEqual(products[0]["section"], "Cruce SAP + Foto")

    def test_16960_photo_is_resolved_by_codigo_dia(self) -> None:
        matches = [product for product in self.products if identifier(product.get("codigoDia")) == "16960"]
        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0].get("photoMatch"), "codigoDia")
        self.assertEqual(matches[0].get("photoFile"), "lote-01/16960.webp")
        self.assertEqual((matches[0].get("visual") or {}).get("src"), "assets/catalog/images/lote-01/16960.webp")
        coverage = json.loads((ROOT / "data/photo-coverage.json").read_text(encoding="utf-8"))
        file_row = next(row for row in coverage["files"] if row.get("identifier") == "16960")
        self.assertEqual(file_row["status"], "matched")
        self.assertEqual(file_row["matchedBy"], "codigoDia")

    def test_16672_photo_is_published_with_cache_revision(self) -> None:
        matches = [product for product in self.products if identifier(product.get("codigoDia")) == "16672"]
        self.assertEqual(len(matches), 1)
        product = matches[0]
        self.assertEqual(product.get("photoMatch"), "codigoDia")
        self.assertEqual(product.get("photoFile"), "lote-01/16672.webp")
        visual = product.get("visual") or {}
        image_path = ROOT / visual.get("src", "")
        self.assertTrue(image_path.is_file())
        self.assertEqual(visual.get("revision"), hashlib.sha256(image_path.read_bytes()).hexdigest()[:12])

    def test_sku_uploads_are_published_only_with_day_code_names(self) -> None:
        matches = [product for product in self.products if identifier(product.get("skuIntl")) == "11186659"]
        self.assertTrue(matches)
        self.assertTrue(all(identifier(product.get("codigoDia")) == "16999" for product in matches))
        self.assertTrue(all(product.get("photoFile") == "lote-01/16999.webp" for product in matches))
        published = [
            path for path in (ROOT / "assets/catalog/images").rglob("*")
            if path.is_file() and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}
        ]
        self.assertTrue(all(path.suffix.lower() == ".webp" and path.stem.isdigit() and len(path.stem) <= 6 for path in published))

    def test_excel_keeps_distinct_sku_intl_and_codigo_dia(self) -> None:
        workbook = load_workbook(ROOT / "Lista_Precios_Base.xlsx", read_only=True, data_only=True)
        sheet = workbook["Base_Campaña"]
        matches = [
            row for row in sheet.iter_rows(min_row=2, values_only=True)
            if identifier(row[0]) == "11160979" and identifier(row[1]) == "16960"
        ]
        workbook.close()
        self.assertEqual(len(matches), 1)
        self.assertNotEqual(identifier(matches[0][0]), identifier(matches[0][1]))

    def test_campaign_rows_without_price_remain_searchable(self) -> None:
        text = (ROOT / "data/products.js").read_text(encoding="utf-8")
        products = json.loads(text.split("window.PRODUCTS = ", 1)[1].split(";\nwindow.PRODUCT_META", 1)[0])
        by_day = {identifier(product.get("codigoDia")): product for product in products}
        for codigo in ("16889", "16972", "16990", "17336", "17337", "17338"):
            self.assertIn(codigo, by_day)
            self.assertEqual(by_day[codigo].get("priceStatus"), "pending")

    def test_every_photo_is_unique_and_crossed_against_all_excel_sources(self) -> None:
        report = json.loads((ROOT / "data/image-source-audit.json").read_text(encoding="utf-8"))
        self.assertEqual(report["status"], "ok")
        self.assertEqual(report["imageFiles"], 98)
        self.assertEqual(report["uniqueCodes"], 98)
        self.assertEqual(report["lotCounts"], {"lote-01": 92, "lote-02": 6, "lote-03": 0, "lote-04": 0})
        self.assertTrue(all(row["excelMatches"] for row in report["files"]))
        workbook_names = {row["workbook"] for row in report["workbooks"]}
        self.assertIn("Lista_Precios_Base.xlsx", workbook_names)
        self.assertTrue(any(name.startswith("engines/merch-lists/") for name in workbook_names))

    def test_day_photo_wins_when_sku_file_has_duplicate_content(self) -> None:
        coverage = json.loads((ROOT / "data/photo-coverage.json").read_text(encoding="utf-8"))
        rows = {row.get("identifier"): row for row in coverage["files"]}
        self.assertEqual(rows["16960"].get("file"), "lote-01/16960.webp")
        sku_duplicate = rows.get("11160979")
        if sku_duplicate is not None:
            self.assertIsNone(sku_duplicate.get("file"))
            self.assertEqual(sku_duplicate.get("status"), "ignored-duplicate-content")
            self.assertEqual(sku_duplicate.get("kept"), "lote-01/16960.jpeg")
        self.assertFalse(any("/11160979." in product.get("photoFile", "") for product in self.products))

    def test_post_publish_audit_rejects_duplicates(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "images"
            for number in range(1, 5):
                (source / f"lote-{number:02d}").mkdir(parents=True, exist_ok=True)
            image = Image.new("RGB", (20, 20), "green")
            image.save(source / "lote-01/16162.jpg")
            image.save(source / "lote-04/16378.jpg")
            with self.assertRaisesRegex(ValueError, "Duplicado detectado después de publicar"):
                validate_published_lots(source)

    def test_interface_contains_catalog_and_quantity_contract(self) -> None:
        html = (ROOT / "index.html").read_text(encoding="utf-8")
        app = (ROOT / "app.js").read_text(encoding="utf-8")
        css = (ROOT / "catalog.css").read_text(encoding="utf-8")
        sw = (ROOT / "sw.js").read_text(encoding="utf-8")
        for token in ("modeMenu", "modeBack", "data-app-mode=\"catalog\"", "data-app-mode=\"merch\"", "data-app-mode=\"export\"", "catalogGrid", "catalogFilters", "catalogPhotoFilter", "catalogSort", "catalogReset", "catalogActiveFilters", "catalogNameToggle", "catalogLoadMore", "woeSearchClear", "microsCatalogResults", "merch-catalog.js"):
            self.assertIn(token, html)
        for token in ("Pendiente de precio", "product-reference-photo", "fotografías disponibles", "missingPhotoWhatsappUrl"):
            self.assertIn(token, app + css)
        self.assertNotIn("catalogLoadAll", html + app)
        self.assertNotIn("microsGroupFilter", html)
        self.assertNotIn("microsFamilyFilter", html)
        self.assertIn("catalogVisibleLimit = 12", app)
        self.assertIn("Diseñado por Jorge Alcantar Aguiar & Enrique César Flores", app)
        self.assertIn("exportStockExcel", app)
        for token in ("isSapPrecountHtml", ".mat-column-lcodSAP", "sunidadMedidaBase", "stockMeta?.preCount", "Inventario previo · Rectificación", "Inventario_Preconteo_Rectificacion"):
            self.assertIn(token, app)
        self.assertIn("stockConfirmExcel", html)
        self.assertIn("Catálogo General", html)
        for token in ("renderCatalog", "selectAppMode", "showHome", "articleKey", "quantity", "Código Día", "Código SAP", "label:'CONTEO'", "'Conteo'"):
            self.assertIn(token, app)
        self.assertNotIn("catalogPrice", app)
        self.assertNotIn("Agregar · $", app)
        self.assertNotIn("catalogVisualDialog", html)
        self.assertNotIn("openCatalogVisual", app)
        self.assertNotIn("data-catalog-visual", app)
        self.assertNotIn("data-woe-visual", app)
        self.assertNotIn("Ampliar", app)
        for token in ("catalogBatchSize", "catalogPhotoState", "catalogSort", "catalogNameMode", "catalogNameSources", "catalogNames", "updateCatalogNameToggle", "codebrew-catalog-state-v2", "stockPriority==='active'", "data-catalog-image", "loading=\"lazy\"", "decoding=\"async\""):
            self.assertIn(token, app)
        for token in ("MICROS_RESULT_BATCH", "microsVisibleLimit", "countFilterActive", "data-micros-load-more", "focusFirstCatalogResult", "e.key==='ArrowDown'"):
            self.assertIn(token, app)
        self.assertNotIn("allMatches.slice(0,40)", app)
        self.assertNotIn("max-height: 420px", css)
        self.assertIn("La búsqueda escrita revisa todo el catálogo", html)
        for token in ("SKU POS", "SKU internacional", "Nombre Inventario", "Descripción SCI"):
            self.assertIn(token, html + app)
        for token in ("grid-template-columns: repeat(5", "object-position: center", ".catalog-options", ".catalog-view-bar"):
            self.assertIn(token, css)
        self.assertIn("Mostrar Descripción SCI", html)
        self.assertNotIn("catalog-card-description", app)
        self.assertIn("missingPhotoWhatsappUrl", app)
        self.assertIn("525521107475", app)
        self.assertIn("https://wa.me/", app)
        self.assertIn("whatsapp://send?phone=", app)
        self.assertIn("data-photo-whatsapp", app)
        self.assertIn("window.location.assign", app)
        self.assertIn("Código Día:", app)
        self.assertIn("Nombre sugerido del archivo:", app)
        self.assertIn("Toma una foto completa y legible del termo", app)
        self.assertIn("Enviar foto por WhatsApp", app)
        self.assertIn("catalog-photo-request", app)
        self.assertIn("catalog-missing-visual", app)
        self.assertIn("photoState=hasPhoto?'':", app)
        self.assertIn("visual?.src))*100000000", app)
        self.assertIn("stockQuantity", app)
        self.assertIn("normalizeSearchText", app)
        self.assertIn("renderWoeSuggestions(input.value)", app)
        self.assertNotIn("input.value='';box.hidden=true", app)
        self.assertIn("scheduleCatalogRender", app)
        self.assertIn("updateViaCache:'none'", app)
        self.assertIn("controllerchange", app)
        self.assertIn("registration.update()", app)
        self.assertNotIn("localStorage.clear", app)
        self.assertNotIn("document.cookie", app)
        for token in ("inputmode=\"search\"", "enterkeyhint=\"search\"", "spellcheck=\"false\""):
            self.assertIn(token, html)
        for token in ("key.startsWith('codebrew-')", "isCoreResource", "freshFirst", "cache:'reload'"):
            self.assertIn(token, sw)
        self.assertIn("visual?.revision", app)
        for token in ("catalog-card-top", "catalog-source", "catalog-match", "Foto disponible", "visualQualityLabel"):
            self.assertNotIn(token, app)

    def test_photo_list_has_one_row_per_codigo_dia(self) -> None:
        with (ROOT / "data/Listado_Codigo_Dia_Fotos.csv").open("r", encoding="utf-8-sig", newline="") as stream:
            rows = list(csv.DictReader(stream))
        codes = {str(product.get("codigoDia")) for product in self.products if product.get("codigoDia")}
        self.assertEqual(len(rows), len(codes))
        self.assertEqual(len({row["CÓDIGO DÍA"] for row in rows}), len(rows))
        self.assertTrue(all(row["CÓDIGO DÍA"] for row in rows))
        self.assertEqual({row["ESTADO FOTO"] for row in rows}.difference({"CON FOTO", "FALTA FOTO"}), set())
        photo_flags = [row["ESTADO FOTO"] == "CON FOTO" for row in rows]
        self.assertEqual(photo_flags, sorted(photo_flags, reverse=True))
        photo_stock = [float(row["EXISTENCIA"] or 0) for row in rows if row["ESTADO FOTO"] == "CON FOTO" and row["PRIORIDAD"] == "ACTIVO"]
        self.assertEqual(photo_stock, sorted(photo_stock, reverse=True))

    def test_stock_priority_and_photo_priority(self) -> None:
        active = [product for product in self.products if product.get("stockPriority") == "active"]
        secondary = [product for product in self.products if product.get("stockPriority") == "secondary"]
        self.assertTrue(active)
        self.assertTrue(secondary)
        self.assertEqual(self.catalog["meta"].get("activeStockProducts"), len(active))
        self.assertEqual(self.catalog["meta"].get("secondaryProducts"), len(secondary))
        first_secondary = next((i for i, product in enumerate(self.products) if product.get("stockPriority") == "secondary"), len(self.products))
        self.assertTrue(all(product.get("stockPriority") == "active" for product in self.products[:first_secondary]))
        active_with_photo = [product for product in active if product.get("visual")]
        self.assertEqual(len(active_with_photo), self.catalog["meta"].get("activeWithPhoto"))
        first_missing = next((index for index, product in enumerate(self.products) if not product.get("visual")), len(self.products))
        self.assertTrue(all(product.get("visual") for product in self.products[:first_missing]))
        photographed = [product for product in self.products if product.get("visual")]
        photographed_quantities = [float(product.get("stockQuantity") or 0) for product in photographed if product.get("stockPriority") == "active"]
        self.assertEqual(photographed_quantities, sorted(photographed_quantities, reverse=True))
        active_missing = [product for product in active if not product.get("visual")]
        quantities = [float(product.get("stockQuantity") or 0) for product in active_missing]
        self.assertEqual(quantities, sorted(quantities, reverse=True))


if __name__ == "__main__":
    unittest.main()
