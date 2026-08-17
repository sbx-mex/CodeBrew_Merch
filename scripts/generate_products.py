#!/usr/bin/env python3
"""Valida Lista_Precios_Base.xlsx y genera los datos públicos de CodeBrew."""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


REQUIRED_SHEETS = ("Base_Campaña", "Discovery", "Homologados", "Essentials")
WOE_SHEETS = ("SAP", "Catalogo Micros")
CAMPAIGN_ALIASES = {
    "si": ("Summer", ("SI", "SII", "Summer")),
    "sii": ("Summer", ("SI", "SII", "Summer")),
    "summer": ("Summer", ("SI", "SII", "Summer")),
    "wc": ("World Cup", ("WC", "World Cup")),
    "world cup": ("World Cup", ("WC", "World Cup")),
    "sp": ("Spring", ("SP", "Spring")),
    "spring": ("Spring", ("SP", "Spring")),
    "wt": ("Winter", ("WT", "Winter")),
    "winter": ("Winter", ("WT", "Winter")),
    "xm": ("Christmas", ("XM", "Christmas")),
    "christmas": ("Christmas", ("XM", "Christmas")),
}
FIELD_ALIASES = {
    "skuIntl": ("sku intl",),
    "codigoDia": ("codigo dia",),
    "descripcion": ("descripcion sci", "descripcion"),
    "nombrePos": ("nombre pos",),
    "nombreInventario": ("nombre inventario",),
    "botonPos": ("boton pos",),
    "skuPos": ("sku pos",),
}
REQUIRED_FIELDS = (
    "skuIntl",
    "codigoDia",
    "descripcion",
    "nombrePos",
    "nombreInventario",
    "skuPos",
)
HEADER_SCAN_LIMIT = 25
SAP_FIELD_ALIASES = {
    "idWoe": ("id woe",),
    "codigoDia": ("codigo dia",),
    "descripcionSap": ("descripcion sap", "descripcion"),
}
MICROS_FIELD_ALIASES = {
    "agrupado": ("agrupado",),
    "familia": ("familia",),
    "conteo": ("conteo",),
    "nombreMicros": ("nombre micros",),
    "codigoDia": ("codigo dia",),
}


def normalize_header(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"\s+", " ", text).casefold()


def clean_text(value: object) -> str:
    return re.sub(r"[ \t]+", " ", str(value or "").strip())


def canonical_text(value: object) -> str:
    """Normaliza espacios sólo para comparar duplicados, sin alterar el dato visible."""
    return re.sub(r"\s+", " ", str(value or "").strip())


def homologate_campaign(value: object):
    """Devuelve campaña homologada sin aplicar coincidencias parciales ambiguas."""
    original = clean_text(value)
    if not original:
        return None
    normalized = normalize_header(original)
    match = CAMPAIGN_ALIASES.get(normalized)
    if not match:
        return None
    name, aliases = match
    return {"original": original, "name": name, "aliases": list(aliases)}


def campaign_from_product_name(value: object):
    """Extrae únicamente un código de campaña completo al inicio del Nombre POS."""
    original = clean_text(value)
    match = re.match(
        r"^(WORLD CUP|CHRISTMAS|SUMMER|SPRING|WINTER|SII|SI|WC|SP|WT|XM)(?=\d|\s|$)",
        original,
        flags=re.IGNORECASE,
    )
    return homologate_campaign(match.group(1)) if match else homologate_campaign(original)


def identifier(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float, Decimal)):
        number = Decimal(str(value))
        if number == number.to_integral():
            text = str(number.quantize(Decimal("1")))
            fmt = str(cell.number_format or "")
            zero_match = re.fullmatch(r"0+", fmt)
            return text.zfill(len(fmt)) if zero_match else text
        return format(number.normalize(), "f")
    return clean_text(value)


def money(value: object) -> str:
    if value is None or clean_text(value) == "":
        return ""
    if isinstance(value, str):
        raw = re.sub(r"[$,\s]", "", value)
    else:
        raw = str(value)
    try:
        amount = Decimal(raw)
    except InvalidOperation:
        raise ValueError(f"precio no numérico: {value!r}")
    return f"${amount:,.2f}"


def locate_named_headers(ws, aliases_by_field, required_fields, *, label: str):
    """Localiza encabezados por nombre sin depender de fila ni orden de columnas."""
    scan_end = min(ws.max_row, HEADER_SCAN_LIMIT)
    best_missing = list(required_fields)
    for row_number in range(1, scan_end + 1):
        cells = next(ws.iter_rows(min_row=row_number, max_row=row_number))
        original = [clean_text(cell.value) for cell in cells]
        normalized = [normalize_header(value) for value in original]
        positions = {}
        duplicate_fields = []
        for field, aliases in aliases_by_field.items():
            matches = [index for index, header in enumerate(normalized) if header in aliases]
            if len(matches) > 1:
                duplicate_fields.append(field)
            elif matches:
                positions[field] = matches[0]
        missing = [field for field in required_fields if field not in positions]
        if len(missing) < len(best_missing):
            best_missing = missing
        if not missing:
            if duplicate_fields:
                raise ValueError(
                    f"{label} fila {row_number}: encabezados duplicados para "
                    f"{', '.join(duplicate_fields)}"
                )
            return row_number, original, normalized, positions
    raise ValueError(
        f"{label}: no se localizaron encabezados válidos en las primeras "
        f"{scan_end} filas; faltan {', '.join(best_missing)}"
    )


def locate_headers(ws):
    header_row, original, normalized, positions = locate_named_headers(
        ws,
        FIELD_ALIASES,
        REQUIRED_FIELDS,
        label=ws.title,
    )
    duplicates = [name for name, count in Counter(normalized).items() if name and count > 1]
    price_columns = {
        original[index].strip().upper(): index
        for index, header in enumerate(normalized)
        if re.fullmatch(r"c[1-6]", header)
    }
    if not price_columns:
        raise ValueError(f"{ws.title} fila {header_row}: no se detectó precio C1-C6")
    return header_row, original, positions, price_columns, duplicates


def sheet_button(sheet_name: str, row, positions) -> str:
    if "botonPos" in positions:
        return clean_text(row[positions["botonPos"]].value)
    if sheet_name == "Base_Campaña":
        return "Campaña"
    if sheet_name == "Essentials":
        return "Essentials"
    return ""


def parse_sheet(ws):
    header_row, headers, positions, price_columns, duplicate_headers = locate_headers(ws)
    products = []
    empty_rows = 0
    invalid_rows = []
    duplicate_rows = 0
    record_rows = 0
    seen = set()

    for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        if all(cell.value is None or clean_text(cell.value) == "" for cell in row):
            empty_rows += 1
            continue
        record_rows += 1
        try:
            nombre_pos = clean_text(row[positions["nombrePos"]].value)
            campaign = campaign_from_product_name(nombre_pos) if ws.title == "Base_Campaña" else None
            product = {
                "skuIntl": identifier(row[positions["skuIntl"]]),
                "codigoDia": identifier(row[positions["codigoDia"]]),
                "descripcion": clean_text(row[positions["descripcion"]].value),
                "nombrePos": nombre_pos,
                "nombreInventario": clean_text(row[positions["nombreInventario"]].value),
                "skuPos": identifier(row[positions["skuPos"]]),
                "botonPos": sheet_button(ws.title, row, positions),
                "tier": {
                    tier: money(row[index].value)
                    for tier, index in price_columns.items()
                    if row[index].value is not None and clean_text(row[index].value) != ""
                },
                "base": "Campaña" if ws.title == "Base_Campaña" else ws.title,
                "sourceSheet": ws.title,
                "sourceRow": row_number,
            }
            if campaign:
                product.update({
                    "campaignOriginal": campaign["original"],
                    "campaign": campaign["name"],
                    "campaignAliases": campaign["aliases"],
                })
        except (IndexError, ValueError) as error:
            invalid_rows.append({"row": row_number, "reason": str(error)})
            continue

        searchable = any(
            product[key] for key in ("skuIntl", "codigoDia", "nombrePos", "nombreInventario", "skuPos")
        )
        if not searchable or not product["tier"]:
            invalid_rows.append({"row": row_number, "reason": "sin identificador útil o sin precio"})
            continue

        fingerprint = json.dumps(
            {key: value for key, value in product.items() if key not in ("sourceRow",)},
            ensure_ascii=False,
            sort_keys=True,
        )
        if fingerprint in seen:
            duplicate_rows += 1
            continue
        seen.add(fingerprint)
        products.append(product)

    report = {
        "sheet": ws.title,
        "headerRow": header_row,
        "headers": headers,
        "records": record_rows,
        "valid": len(products),
        "emptyRows": empty_rows,
        "duplicatesIgnored": duplicate_rows,
        "invalidRows": invalid_rows,
        "duplicateHeaders": duplicate_headers,
        "campaigns": dict(Counter(
            product["campaign"] for product in products if product.get("campaign")
        )),
    }
    return products, report


def parse_woe_catalog(workbook, products):
    """Construye el catálogo WOE sin perder relaciones uno-a-varios."""
    missing = [name for name in WOE_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise ValueError(f"faltan pestañas WOE: {', '.join(missing)}")

    sap_ws = workbook["SAP"]
    micros_ws = workbook["Catalogo Micros"]
    sap_header_row, _, _, sap_positions = locate_named_headers(
        sap_ws,
        SAP_FIELD_ALIASES,
        tuple(SAP_FIELD_ALIASES),
        label="SAP",
    )
    micros_header_row, _, _, micros_positions = locate_named_headers(
        micros_ws,
        MICROS_FIELD_ALIASES,
        tuple(MICROS_FIELD_ALIASES),
        label="Catalogo Micros",
    )

    micros_by_dia = {}
    micros_meta_by_dia = {}
    micros_rows = 0
    invalid_micros_rows = []
    for row_number, row in enumerate(
        micros_ws.iter_rows(min_row=micros_header_row + 1),
        start=micros_header_row + 1,
    ):
        code = identifier(row[micros_positions["codigoDia"]])
        name = clean_text(row[micros_positions["nombreMicros"]].value)
        agrupado = clean_text(row[micros_positions["agrupado"]].value)
        familia = clean_text(row[micros_positions["familia"]].value)
        conteo = clean_text(row[micros_positions["conteo"]].value)
        if not code and not name and not agrupado and not familia and not conteo:
            continue
        micros_rows += 1
        if not code:
            invalid_micros_rows.append({"row": row_number, "reason": "falta Codigo DIA"})
            continue
        values = micros_by_dia.setdefault(code, [])
        if name and name not in values:
            values.append(name)
        meta = micros_meta_by_dia.setdefault(code, [])
        detail = {"agrupado": agrupado, "familia": familia, "conteo": conteo, "nombre": name}
        if detail not in meta:
            meta.append(detail)

    merch_by_dia = {}
    for product in products:
        code = clean_text(product.get("codigoDia"))
        if not code:
            continue
        compact = {
            "base": product.get("base", ""),
            "descripcionSci": product.get("descripcion", ""),
            "nombrePos": product.get("nombrePos", ""),
            "nombreInventario": product.get("nombreInventario", ""),
            "skuIntl": product.get("skuIntl", ""),
            "skuPos": product.get("skuPos", ""),
        }
        rows = merch_by_dia.setdefault(code, [])
        if compact not in rows:
            rows.append(compact)

    catalog = []
    seen_woe = Counter()
    seen_dia = Counter()
    sap_rows = 0
    exact_sap_duplicates = 0
    invalid_sap_rows = []
    seen_sap_rows = set()
    sap_codes = set()
    for row_number, row in enumerate(
        sap_ws.iter_rows(min_row=sap_header_row + 1),
        start=sap_header_row + 1,
    ):
        woe_id = identifier(row[sap_positions["idWoe"]])
        code = identifier(row[sap_positions["codigoDia"]])
        description = clean_text(row[sap_positions["descripcionSap"]].value)
        if not woe_id and not code and not description:
            continue
        sap_rows += 1
        if not woe_id or not code:
            missing = []
            if not woe_id:
                missing.append("ID WOE")
            if not code:
                missing.append("Codigo DIA")
            invalid_sap_rows.append({"row": row_number, "reason": f"falta {' y '.join(missing)}"})
            continue
        fingerprint = (woe_id, code, canonical_text(description))
        if fingerprint in seen_sap_rows:
            exact_sap_duplicates += 1
            continue
        seen_sap_rows.add(fingerprint)
        sap_codes.add(code)
        seen_woe[woe_id] += 1
        seen_dia[code] += 1
        micros = micros_by_dia.get(code, [])
        merch = merch_by_dia.get(code, [])
        catalog.append({
            "idWoe": woe_id,
            "codigoDia": code,
            "descripcionSap": description,
            "micros": micros,
            "microsMeta": micros_meta_by_dia.get(code, []),
            "agrupado": sorted({item["agrupado"] for item in micros_meta_by_dia.get(code, []) if item["agrupado"]}),
            "familia": sorted({item["familia"] for item in micros_meta_by_dia.get(code, []) if item["familia"]}),
            "conteo": sorted({item["conteo"] for item in micros_meta_by_dia.get(code, []) if item["conteo"]}),
            "unidadMicros": "",
            "merch": merch,
            "validation": {
                "sap": bool(description),
                "micros": bool(micros),
                "merch": bool(merch),
            },
            "operationalValidation": {
                "sapMicros": bool(description and micros),
                "merchRequired": False,
            },
            "sourceRow": row_number,
            "origin": "SAP",
        })

    orphan_codes = sorted(
        (set(micros_by_dia) | set(merch_by_dia)) - sap_codes,
        key=lambda value: (not value.isdigit(), int(value) if value.isdigit() else value),
    )
    for code in orphan_codes:
        micros = micros_by_dia.get(code, [])
        merch = merch_by_dia.get(code, [])
        catalog.append({
            "idWoe": "",
            "codigoDia": code,
            "descripcionSap": "",
            "micros": micros,
            "microsMeta": micros_meta_by_dia.get(code, []),
            "agrupado": sorted({item["agrupado"] for item in micros_meta_by_dia.get(code, []) if item["agrupado"]}),
            "familia": sorted({item["familia"] for item in micros_meta_by_dia.get(code, []) if item["familia"]}),
            "conteo": sorted({item["conteo"] for item in micros_meta_by_dia.get(code, []) if item["conteo"]}),
            "unidadMicros": "",
            "merch": merch,
            "validation": {"sap": False, "micros": bool(micros), "merch": bool(merch)},
            "operationalValidation": {"sapMicros": False, "merchRequired": False},
            "sourceRow": None,
            "origin": "Micros/MERCH",
        })

    if not catalog:
        raise ValueError("SAP no contiene registros válidos")

    report = {
        "sapRows": sap_rows,
        "sapHeaderRow": sap_header_row,
        "catalogRows": len(catalog),
        "sapCatalogRows": len(seen_sap_rows),
        "orphanCatalogRows": len(orphan_codes),
        "uniqueWoe": len(seen_woe),
        "uniqueDiaSap": len(seen_dia),
        "exactSapDuplicatesIgnored": exact_sap_duplicates,
        "invalidSapRows": invalid_sap_rows,
        "multiWoeRelations": sum(count - 1 for count in seen_woe.values()),
        "multiDiaRelations": sum(count - 1 for count in seen_dia.values()),
        "microsRows": micros_rows,
        "microsHeaderRow": micros_header_row,
        "invalidMicrosRows": invalid_micros_rows,
        "uniqueDiaMicros": len(micros_by_dia),
        "microsUnitsPopulated": 0,
        "microsGroups": len({item["agrupado"] for rows in micros_meta_by_dia.values() for item in rows if item["agrupado"]}),
        "microsFamilies": len({item["familia"] for rows in micros_meta_by_dia.values() for item in rows if item["familia"]}),
        "microsCountGroups": len({item["conteo"] for rows in micros_meta_by_dia.values() for item in rows if item["conteo"]}),
        "withMicros": sum(1 for item in catalog if item["validation"]["micros"]),
        "withMerch": sum(1 for item in catalog if item["validation"]["merch"]),
        "withoutSap": sum(1 for item in catalog if not item["validation"]["sap"]),
        "operationalSapMicrosMatch": sum(1 for item in catalog if item["operationalValidation"]["sapMicros"]),
        "operationalNeedsReview": sum(1 for item in catalog if not item["operationalValidation"]["sapMicros"]),
        "completeTripleMatch": sum(1 for item in catalog if item["validation"]["sap"] and item["validation"]["micros"] and item["validation"]["merch"]),
    }
    return catalog, report


def generate(excel_path: Path, js_path: Path, woe_path: Path, report_path: Path):
    workbook = load_workbook(excel_path, data_only=True, read_only=False)
    missing_sheets = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(f"faltan pestañas obligatorias: {', '.join(missing_sheets)}")

    products = []
    reports = []
    for sheet_name in REQUIRED_SHEETS:
        parsed, report = parse_sheet(workbook[sheet_name])
        products.extend(parsed)
        reports.append(report)

    woe_catalog, woe_report = parse_woe_catalog(workbook, products)

    campaign = workbook["Base_Campaña"]
    _, _, campaign_positions, _, _ = locate_headers(campaign)
    campaign_products = sorted(
        (product for product in products if product["sourceSheet"] == "Base_Campaña"),
        key=lambda product: product["sourceRow"],
    )
    if not campaign_products:
        raise ValueError("Base_Campaña no contiene artículos válidos")
    latest_product = campaign_products[0]
    latest_item = latest_product["nombrePos"]
    latest_item_cell = (
        f"Base_Campaña!{get_column_letter(campaign_positions['nombrePos'] + 1)}"
        f"{latest_product['sourceRow']}"
    )

    # Algunos libros no incluyen docProps/core.xml. En ese caso openpyxl crea
    # fechas con la hora actual al abrirlos, lo que vuelve distinto cada build.
    # El mtime del archivo motor es estable y permite salidas reproducibles.
    generated_at = datetime.fromtimestamp(
        excel_path.stat().st_mtime,
        tz=timezone.utc,
    ).isoformat(timespec="seconds")
    meta = {
        "sourceFile": excel_path.name,
        "generatedAtUtc": generated_at,
        "latestItem": latest_item,
        "totalProducts": len(products),
        "sheets": {item["sheet"]: item["valid"] for item in reports},
        "woeRecords": len(woe_catalog),
    }
    report = {
        "status": "ok",
        "sourceFile": excel_path.name,
        "generatedAtUtc": generated_at,
        "sourceModifiedAtUtc": generated_at,
        "latestItemCell": latest_item_cell,
        "latestItem": latest_item,
        "totalValidProducts": len(products),
        "sheets": reports,
        "woe": woe_report,
    }

    js_path.parent.mkdir(parents=True, exist_ok=True)
    woe_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    js_path.write_text(
        "window.PRODUCTS = "
        + json.dumps(products, ensure_ascii=False, separators=(",", ":"))
        + ";\nwindow.PRODUCT_META = "
        + json.dumps(meta, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    woe_path.write_text(
        "window.WOE_CATALOG = "
        + json.dumps(woe_catalog, ensure_ascii=False, separators=(",", ":"))
        + ";\nwindow.WOE_META = "
        + json.dumps(woe_report, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default="Lista_Precios_Base.xlsx")
    parser.add_argument("--output", default="data/products.js")
    parser.add_argument("--woe-output", default="data/woe.js")
    parser.add_argument("--report", default="data/import-report.json")
    args = parser.parse_args()
    try:
        generate(Path(args.excel), Path(args.output), Path(args.woe_output), Path(args.report))
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
