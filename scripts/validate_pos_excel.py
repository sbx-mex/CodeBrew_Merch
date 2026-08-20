#!/usr/bin/env python3
"""Valida que SKU POS provenga del Excel y no de parches de ejecución."""
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from generate_products import REQUIRED_SHEETS, clean_text, identifier, locate_headers


def load_products(path: Path) -> list[dict]:
    text = path.read_text(encoding="utf-8")
    match = re.search(r"window\.PRODUCTS\s*=\s*(\[.*?\]);\s*window\.", text, flags=re.S)
    if not match:
        match = re.search(r"window\.PRODUCTS\s*=\s*(\[.*\])\s*;?\s*$", text, flags=re.S)
    if not match:
        raise ValueError(f"No se encontró window.PRODUCTS en {path}")
    data = json.loads(match.group(1))
    if not isinstance(data, list):
        raise ValueError("window.PRODUCTS no es una lista")
    return data


def excel_pos_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    for sheet_name in REQUIRED_SHEETS:
        ws = wb[sheet_name]
        header_row, _, positions, _, _ = locate_headers(ws)
        for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
            if all(cell.value is None or clean_text(cell.value) == "" for cell in row):
                continue
            sku_pos = identifier(row[positions["skuPos"]])
            codigo_dia = identifier(row[positions["codigoDia"]])
            sku_intl = identifier(row[positions["skuIntl"]])
            nombre_pos = clean_text(row[positions["nombrePos"]].value)
            nombre_inv = clean_text(row[positions["nombreInventario"]].value)
            if not any((sku_pos, codigo_dia, sku_intl, nombre_pos, nombre_inv)):
                continue
            rows.append({
                "sourceSheet": sheet_name,
                "sourceRow": row_number,
                "skuPos": sku_pos,
                "codigoDia": codigo_dia,
                "skuIntl": sku_intl,
                "nombrePos": nombre_pos,
                "nombreInventario": nombre_inv,
            })
    return rows


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", type=Path, default=Path("Lista_Precios_Base.xlsx"))
    parser.add_argument("--products", type=Path, default=Path("data/products.js"))
    parser.add_argument("--project", type=Path, default=Path("."))
    parser.add_argument("--report", type=Path, default=Path("data/pos-excel-validation.json"))
    args = parser.parse_args()

    root = args.project.resolve()
    excel_rows = excel_pos_rows(args.excel.resolve())
    products = load_products(args.products.resolve())
    generated = {(str(p.get("sourceSheet") or ""), int(p.get("sourceRow") or 0)): p for p in products}

    mismatches = []
    missing = []
    for row in excel_rows:
        key = (row["sourceSheet"], row["sourceRow"])
        product = generated.get(key)
        if not product:
            # Sólo exige fila publicada cuando tiene precio y fue aceptada por el generador.
            continue
        if str(product.get("skuPos") or "") != row["skuPos"]:
            mismatches.append({"key": key, "excel": row["skuPos"], "generated": product.get("skuPos", "")})
        if row["skuPos"] and not str(product.get("skuPos") or ""):
            missing.append({"key": key, "excel": row["skuPos"]})

    forbidden_tokens = ("POS_MIRROR", "SKU POS · ESPEJO", "pos-operational-overrides.js")
    runtime_files = [root / "app.js", root / "index.html", root / "sw.js"]
    forbidden = []
    for path in runtime_files:
        text = path.read_text(encoding="utf-8")
        for token in forbidden_tokens:
            if token in text:
                forbidden.append({"file": path.relative_to(root).as_posix(), "token": token})

    by_pos = defaultdict(list)
    for product in products:
        sku = str(product.get("skuPos") or "").strip()
        if sku:
            by_pos[sku].append({"sheet": product.get("sourceSheet"), "row": product.get("sourceRow"), "name": product.get("nombrePos")})

    duplicate_groups = {sku: refs for sku, refs in by_pos.items() if len(refs) > 1}
    report = {
        "ok": not mismatches and not missing and not forbidden,
        "source": args.excel.name,
        "excelRowsReviewed": len(excel_rows),
        "generatedProductsReviewed": len(products),
        "mismatches": mismatches,
        "missingSkuPos": missing,
        "forbiddenRuntimeOverrides": forbidden,
        "duplicateSkuPosSummary": {
            "groups": len(duplicate_groups),
            "rows": sum(len(refs) for refs in duplicate_groups.values()),
            "sample": [
                {"skuPos": sku, "count": len(refs), "examples": refs[:3]}
                for sku, refs in list(duplicate_groups.items())[:10]
            ],
            "blocking": False,
        },
        "rule": "SKU POS se toma directamente del encabezado SKU POS del Excel; filas y orden pueden cambiar sin usar parches por artículo.",
    }
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
