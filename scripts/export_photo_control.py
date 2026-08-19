#!/usr/bin/env python3
"""Exporta un renglón por Código Día y destaca los homologados."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = (
    "ESTADO FOTO", "EXISTENCIA", "CÓDIGO DÍA", "ID SAP", "SKU INTL", "ARTÍCULO",
    "DESCRIPCIÓN SAP", "HOMOLOGADO", "PRIORIDAD", "COINCIDENCIA STOCK", "FOTO", "LOTE",
    "COINCIDENCIA FOTO",
)
GREEN = "006241"
DARK_GREEN = "073F2F"
LIGHT_GREEN = "DCF3E8"
LIGHT_RED = "FFE6E2"


def load_catalog(path: Path) -> dict:
    text = path.read_text(encoding="utf-8")
    return json.loads(text.split("=", 1)[1].strip().rstrip(";"))


def numeric_code(value: object) -> tuple[int, str]:
    text = str(value or "")
    return (int(text), text) if text.isdigit() else (999_999_999, text.casefold())


def control_rows(payload: dict) -> list[list[object]]:
    groups: dict[str, list[dict]] = {}
    for product in payload.get("products", []):
        code = str(product.get("codigoDia") or "").strip()
        if code:
            groups.setdefault(code, []).append(product)
    rows: list[list[object]] = []
    for code, products in groups.items():
        products.sort(key=lambda product: (
            not bool(product.get("visual")),
            product.get("stockPriority") != "active",
            product.get("source") != "Genéricos homologados",
        ))
        representative = products[0]
        photos = sorted({str(product.get("photoFile")) for product in products if product.get("photoFile")})
        skus = sorted({str(product.get("skuIntl")) for product in products if str(product.get("skuIntl") or "") not in {"", "-", "N/A"}})
        matches = sorted({str(product.get("photoMatch")) for product in products if product.get("photoMatch")})
        stock_matches = sorted({str(product.get("stockMatchName")) for product in products if product.get("stockMatchName")})
        stock_quantity = max(float(product.get("stockQuantity") or 0) for product in products)
        sap_ids = sorted({str(value) for product in products for value in product.get("sapIds", []) if str(value)})
        sap_descriptions = sorted({str(value) for product in products for value in product.get("sapDescriptions", []) if str(value)})
        rows.append([
            "CON FOTO" if photos else "FALTA FOTO",
            int(stock_quantity) if stock_quantity.is_integer() else stock_quantity,
            code,
            ", ".join(sap_ids),
            ", ".join(skus),
            representative.get("displayName") or representative.get("descripcionSci") or "Artículo",
            " | ".join(sap_descriptions),
            "SÍ" if any(product.get("source") == "Genéricos homologados" for product in products) else "NO",
            "ACTIVO" if any(product.get("stockPriority") == "active" for product in products) else "SECUNDARIO",
            ", ".join(stock_matches),
            ", ".join(photos),
            ", ".join(sorted({photo.split("/", 1)[0] for photo in photos})),
            ", ".join(matches),
        ])
    rows.sort(key=lambda row: (
        row[0] != "CON FOTO",
        row[8] != "ACTIVO",
        -float(row[1] or 0),
        not bool(row[3]),
        row[7] != "SÍ",
        numeric_code(row[2]),
    ))
    return rows


def write_csv(path: Path, rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream, lineterminator="\n")
        writer.writerow(HEADERS)
        writer.writerows(rows)


def write_xlsx(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Control de fotos"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:M1")
    sheet["A1"] = "CONTROL DE FOTOS · DISPONIBLES PRIMERO"
    sheet["A1"].fill = PatternFill("solid", fgColor=GREEN)
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 34

    sheet["A3"], sheet["C3"], sheet["E3"], sheet["G3"] = "CÓDIGOS DÍA", "CON FOTO", "FALTAN", "COBERTURA"
    last_row = len(rows) + 7
    sheet["A4"] = f"=COUNTA(C8:C{last_row})"
    sheet["C4"] = f'=COUNTIF(A8:A{last_row},"CON FOTO")'
    sheet["E4"] = f'=COUNTIF(A8:A{last_row},"FALTA FOTO")'
    sheet["G4"] = "=IFERROR(C4/A4,0)"
    sheet["G4"].number_format = "0.0%"
    for start in ("A3:B4", "C3:D4", "E3:F4", "G3:H4"):
        for row in sheet[start]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="EAF4EF")
                cell.font = Font(color=DARK_GREEN, bold=True, size=14 if cell.row == 4 else 11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=Side(style="thin", color="BFD6CC"))

    sheet.merge_cells("A6:M6")
    sheet["A6"] = "Primero aparecen los artículos CON FOTO, ordenados por existencia de mayor a menor. Después se muestran claramente los pendientes como FALTA FOTO."
    sheet["A6"].fill = PatternFill("solid", fgColor="FFF3D6")
    sheet["A6"].font = Font(color="6A4A00", italic=True)
    sheet["A6"].alignment = Alignment(wrap_text=True)
    sheet.append(list(HEADERS))
    for row in rows:
        sheet.append(row)

    for cell in sheet[7]:
        cell.fill = PatternFill("solid", fgColor=DARK_GREEN)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if rows:
        table = Table(displayName="ControlFotosCodigoDia", ref=f"A7:M{last_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        sheet.add_table(table)
        sheet.conditional_formatting.add(f"A8:A{last_row}", FormulaRule(formula=['A8="CON FOTO"'], fill=PatternFill("solid", fgColor=LIGHT_GREEN), font=Font(color=GREEN, bold=True)))
        sheet.conditional_formatting.add(f"A8:A{last_row}", FormulaRule(formula=['A8="FALTA FOTO"'], fill=PatternFill("solid", fgColor=LIGHT_RED), font=Font(color="A53A2A", bold=True)))
    for row in range(8, last_row + 1):
        sheet.cell(row, 2).number_format = "#,##0.##"
        for column in (3, 4, 5):
            sheet.cell(row, column).number_format = "@"
        sheet.row_dimensions[row].height = 24
    widths = {"A": 15, "B": 14, "C": 13, "D": 16, "E": 18, "F": 30, "G": 42, "H": 13, "I": 13, "J": 28, "K": 32, "L": 12, "M": 18}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A8"
    sheet.auto_filter.ref = f"A7:M{last_row}"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(".tmp.xlsx")
    workbook.save(temporary)
    load_workbook(temporary, read_only=True).close()
    temporary.replace(path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--catalog", type=Path, default=Path("data/merch-catalog.js"))
    parser.add_argument("--xlsx", type=Path, default=Path("Control_Fotos_CodeBrew.xlsx"))
    parser.add_argument("--csv", type=Path, default=Path("data/Listado_Codigo_Dia_Fotos.csv"))
    args = parser.parse_args()
    rows = control_rows(load_catalog(args.catalog))
    write_csv(args.csv, rows)
    write_xlsx(args.xlsx, rows)
    with_photo = sum(row[0] == "CON FOTO" for row in rows)
    print(json.dumps({"codigosDia": len(rows), "conFoto": with_photo, "faltan": len(rows) - with_photo}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
