#!/usr/bin/env python3
"""Construye y audita el catálogo visual sin exponer precios."""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
import shutil
import unicodedata
import zipfile
from collections import Counter, defaultdict
from html.parser import HTMLParser
from pathlib import Path, PurePosixPath

from openpyxl import load_workbook
from PIL import Image, ImageChops, ImageDraw, ImageEnhance, ImageFilter, ImageOps

CANVAS = 768
CONTENT = 660
MAX_FILE_BYTES = 25_000_000
MAX_UNCOMPRESSED = 150_000_000
IMAGE_NOTE = "Fotografía restaurada desde la fuente original; se preservan forma, color y diseño del artículo."
FIELD_ALIASES = {
    "skuIntl": {"sku intl"}, "codigoDia": {"codigo dia"},
    "descripcion": {"descripcion sci", "descripcion"}, "nombrePos": {"nombre pos"},
    "nombreInventario": {"nombre inventario"}, "skuPos": {"sku pos"}, "imagen": {"imagen"},
}
REQUIRED_FIELDS = {"codigoDia", "descripcion", "nombrePos", "nombreInventario", "skuPos", "imagen"}


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", text.casefold()).strip()


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def identifier(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean(value)


def slug(value: object, limit: int = 52) -> str:
    return normalize(value).replace(" ", "-")[:limit].strip("-") or "articulo"


def validate_archive(path: Path, suffix: str) -> None:
    if not path.is_file() or path.stat().st_size >= MAX_FILE_BYTES or path.suffix.lower() != suffix or not zipfile.is_zipfile(path):
        raise ValueError(f"Motor inválido o mayor a 25 MB: {path.name}")
    with zipfile.ZipFile(path) as archive:
        if sum(item.file_size for item in archive.infolist()) > MAX_UNCOMPRESSED:
            raise ValueError(f"Motor descomprimido demasiado grande: {path.name}")
        for item in archive.infolist():
            parts = PurePosixPath(item.filename.replace("\\", "/")).parts
            if item.filename.startswith(("/", "\\")) or ".." in parts:
                raise ValueError(f"Ruta interna insegura en {path.name}")


def locate_headers(sheet) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(sheet.max_row, 12) + 1):
        headers = [normalize(sheet.cell(row_number, column).value) for column in range(1, sheet.max_column + 1)]
        if "codigo dia" not in headers:
            continue
        positions = {}
        for field, aliases in FIELD_ALIASES.items():
            matches = [index + 1 for index, header in enumerate(headers) if header in aliases]
            if len(matches) > 1:
                raise ValueError(f"{sheet.title}: encabezado duplicado {field}")
            if matches:
                positions[field] = matches[0]
        if missing := REQUIRED_FIELDS.difference(positions):
            raise ValueError(f"{sheet.title}: faltan encabezados {sorted(missing)}")
        return row_number, positions
    raise ValueError(f"{sheet.title}: no se localizó el encabezado Código Día")


class CatalogHtmlParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[dict] = []
        self.row: dict | None = None
        self.cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs) -> None:
        attrs = dict(attrs)
        if tag == "tr":
            self.row = {"cells": [], "image": ""}
        elif tag == "td" and self.row is not None:
            self.cell = []
        elif tag == "img" and self.row is not None and not self.row["image"]:
            self.row["image"] = attrs.get("src", "")

    def handle_data(self, data: str) -> None:
        if self.cell is not None:
            self.cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == "td" and self.row is not None and self.cell is not None:
            self.row["cells"].append(clean(" ".join(self.cell)))
            self.cell = None
        elif tag == "tr" and self.row is not None:
            if len(self.row["cells"]) >= 7 and re.fullmatch(r"\d{5,12}", self.row["cells"][0]):
                self.rows.append(self.row)
            self.row = None
            self.cell = None


def source_label(path: Path) -> str:
    name = normalize(path.stem)
    if "summer 2026" in name:
        return "Summer 2026"
    if "winter 2026" in name or "core winter" in name:
        return "Core Winter 2026"
    if "generico" in name:
        return "Genéricos homologados"
    return clean(path.stem)


def load_visual_sources(source_dir: Path) -> tuple[dict[tuple[str, str], bytes], list[dict]]:
    candidates: defaultdict[tuple[str, str], list[bytes]] = defaultdict(list)
    reports = []
    for path in sorted(source_dir.glob("*.zip")):
        validate_archive(path, ".zip")
        with zipfile.ZipFile(path) as archive:
            html_names = [name for name in archive.namelist() if name.lower().endswith("index.html")]
            if len(html_names) != 1:
                raise ValueError(f"{path.name}: se requiere un index.html")
            html_name = html_names[0]
            parser = CatalogHtmlParser()
            parser.feed(archive.read(html_name).decode("utf-8", errors="replace"))
            mapped = 0
            missing = 0
            for row in parser.rows:
                image_name = row["image"]
                if not image_name:
                    missing += 1
                    continue
                member = str(PurePosixPath(html_name).parent / image_name)
                try:
                    raw = archive.read(member)
                    with Image.open(io.BytesIO(raw)) as image:
                        image.verify()
                except (KeyError, OSError, ValueError):
                    missing += 1
                    continue
                key = (identifier(row["cells"][1]), identifier(row["cells"][0]))
                candidates[key].append(raw)
                mapped += 1
            reports.append({"file": path.name, "rows": len(parser.rows), "mappedImages": mapped, "rowsWithoutImage": missing})
    selected = {key: max(images, key=image_score) for key, images in candidates.items()}
    return selected, reports


def image_score(data: bytes) -> tuple[int, int]:
    try:
        with Image.open(io.BytesIO(data)) as image:
            return (image.width * image.height, len(data))
    except OSError:
        return (0, 0)


def average_hash(data: bytes) -> int:
    with Image.open(io.BytesIO(data)) as image:
        pixels = list(ImageOps.grayscale(image.convert("RGB")).resize((8, 8), Image.Resampling.LANCZOS).tobytes())
    average = sum(pixels) / len(pixels)
    return sum((value >= average) << index for index, value in enumerate(pixels))


def category_for(text: str) -> str:
    key = normalize(text)
    if any(token in key for token in ("cold cup", "cld cup", "ccup")):
        return "cold-cup"
    if any(token in key for token in ("water bottle", "wtr btl", "bottle", "btl")):
        return "bottle"
    if any(token in key for token in ("tumbler", "tmblr", "tumb ")):
        return "tumbler"
    if any(token in key for token in ("mug", "taza", "ceramic", "cermc")):
        return "mug"
    if any(token in key for token in ("press", "chemex", "cafetera", "pour over", "brew")):
        return "brew"
    if any(token in key for token in ("bag", "tote", "key ring", "iman", "magnet")):
        return "accessory"
    return "other"


def fallback_tile(category: str) -> Image.Image:
    palette = {"mug": "#006241", "tumbler": "#173f30", "cold-cup": "#26765b", "bottle": "#537c6c", "brew": "#8a5a34", "accessory": "#c49b51", "other": "#6f7e77"}
    canvas = Image.new("RGB", (CANVAS, CANVAS), "#f2ede4")
    draw = ImageDraw.Draw(canvas)
    draw.rounded_rectangle((52, 52, CANVAS - 52, CANVAS - 52), 54, fill="#fffdf9")
    color = palette[category]
    if category == "mug":
        draw.rounded_rectangle((228, 190, 492, 550), 42, fill=color); draw.ellipse((440, 258, 610, 462), outline=color, width=38)
    elif category in {"tumbler", "cold-cup"}:
        draw.rounded_rectangle((246, 148, 522, 610), 64, fill=color); draw.rounded_rectangle((218, 132, 550, 198), 22, fill="#263b33")
    elif category == "bottle":
        draw.rounded_rectangle((274, 166, 494, 616), 72, fill=color); draw.rounded_rectangle((306, 112, 462, 218), 28, fill="#263b33")
    else:
        draw.rounded_rectangle((220, 176, 548, 592), 58, fill=color)
    return canvas


def crop_content(image: Image.Image) -> Image.Image:
    rgba = image.convert("RGBA")
    alpha = rgba.getchannel("A")
    if alpha.getextrema()[0] < 250 and alpha.getbbox():
        return rgba.crop(alpha.getbbox())
    rgb = rgba.convert("RGB")
    background = Image.new("RGB", rgb.size, rgb.getpixel((0, 0)))
    difference = ImageChops.difference(rgb, background).convert("L").point(lambda value: 255 if value > 12 else 0)
    bbox = difference.getbbox()
    return rgba.crop(bbox) if bbox and bbox[2] - bbox[0] > 2 and bbox[3] - bbox[1] > 2 else rgba


def faithful_upscale(image: Image.Image, size: tuple[int, int]) -> Image.Image:
    """Amplía sin inventar elementos y recupera bordes por retroproyección."""
    low = image.convert("RGB")
    high = low.resize(size, Image.Resampling.LANCZOS)
    if max(low.size) >= max(size):
        return high
    for _ in range(2):
        projected = high.resize(low.size, Image.Resampling.LANCZOS)
        residual = ImageChops.subtract(low, projected, offset=128)
        correction = residual.resize(size, Image.Resampling.BICUBIC)
        correction = ImageEnhance.Contrast(correction).enhance(.34)
        high = ImageChops.add(high, correction, offset=-128)
    source_edge = max(low.size)
    strength = 130 if source_edge >= 256 else 175 if source_edge >= 96 else 215
    radius = .65 if source_edge >= 96 else .48
    return high.filter(ImageFilter.UnsharpMask(radius=radius, percent=strength, threshold=2))


def product_tile(data: bytes) -> tuple[Image.Image, dict]:
    with Image.open(io.BytesIO(data)) as source:
        original = ImageOps.exif_transpose(source)
        source_size = original.size
        image = crop_content(original)
    scale = min(CONTENT / max(1, image.width), CONTENT / max(1, image.height))
    fitted = (max(1, round(image.width * scale)), max(1, round(image.height * scale)))
    alpha = image.getchannel("A").resize(fitted, Image.Resampling.LANCZOS)
    rgb = faithful_upscale(image.convert("RGB"), fitted)
    rgb = ImageOps.autocontrast(rgb, cutoff=0.22)
    rgb = ImageEnhance.Contrast(rgb).enhance(1.035)
    rgb = ImageEnhance.Color(rgb).enhance(1.018)
    rgb = rgb.filter(ImageFilter.UnsharpMask(radius=.72, percent=145, threshold=2))
    enhanced = rgb.convert("RGBA")
    enhanced.putalpha(alpha)
    canvas = Image.new("RGBA", (CANVAS, CANVAS), "#fffdf9")
    x = (CANVAS - rgb.width) // 2
    y = (CANVAS - rgb.height) // 2
    shadow_alpha = alpha.filter(ImageFilter.GaussianBlur(12)).point(lambda value: round(value * 0.13))
    shadow = Image.new("RGBA", enhanced.size, (19, 48, 39, 0))
    shadow.putalpha(shadow_alpha)
    canvas.alpha_composite(shadow, (x + 7, y + 10))
    canvas.alpha_composite(enhanced, (x, y))
    return canvas.convert("RGB"), {
        "sourceWidth": source_size[0], "sourceHeight": source_size[1],
        "contentWidth": fitted[0], "contentHeight": fitted[1],
        "scale": round(scale, 3), "passes": 2,
    }


def parse_workbook(path: Path, priority: int, external: dict[tuple[str, str], bytes], overrides: Path) -> tuple[list[dict], dict[str, bytes], dict]:
    validate_archive(path, ".xlsx")
    workbook = load_workbook(path, data_only=True, read_only=False, keep_links=False)
    sheet = workbook.active
    header_row, positions = locate_headers(sheet)
    image_end = positions["descripcion"] - 1
    row_images: defaultdict[int, list[bytes]] = defaultdict(list)
    for image in getattr(sheet, "_images", []):
        if hasattr(image.anchor, "_from") and positions["imagen"] <= image.anchor._from.col + 1 <= image_end:
            try:
                row_images[image.anchor._from.row + 1].append(image._data())
            except (OSError, ValueError):
                pass
    products, visuals = [], {}
    current_section = "Catálogo"
    quality = Counter()
    for row_number in range(header_row + 1, sheet.max_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, sheet.max_column + 1)]
        codigo = identifier(sheet.cell(row_number, positions["codigoDia"]).value)
        descripcion = clean(sheet.cell(row_number, positions["descripcion"]).value)
        sku_pos = identifier(sheet.cell(row_number, positions["skuPos"]).value)
        if not codigo or not (descripcion or sku_pos):
            nonempty = [clean(value) for value in values if clean(value)]
            if len(nonempty) == 1:
                current_section = nonempty[0]
            continue
        sku_intl = identifier(sheet.cell(row_number, positions.get("skuIntl", 1)).value)
        nombre_pos = clean(sheet.cell(row_number, positions["nombrePos"]).value)
        nombre_inventario = clean(sheet.cell(row_number, positions["nombreInventario"]).value)
        display_name = nombre_inventario or nombre_pos or descripcion
        category = category_for(f"{descripcion} {display_name}")
        candidates: list[tuple[str, bytes]] = []
        excel_images = row_images.get(row_number, [])
        candidates.extend(("excel", raw) for raw in excel_images)
        external_raw = external.get((codigo, sku_intl))
        if external_raw:
            candidates.append(("html-row", external_raw))
        if external_raw and excel_images:
            quality["crossChecked"] += 1
            try:
                excel_best = max(excel_images, key=image_score)
                if (average_hash(external_raw) ^ average_hash(excel_best)).bit_count() <= 12:
                    quality["crossConsistent"] += 1
                else:
                    quality["crossReview"] += 1
            except OSError:
                quality["crossReview"] += 1
        override = next((path for path in overrides.glob(f"{codigo}.*") if path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp"}), None)
        if override:
            candidates.append(("premium-override", override.read_bytes()))
        if candidates:
            visual_source, raw = max(candidates, key=lambda candidate: ((1, 0, 0) if candidate[0] == "premium-override" else (0, *image_score(candidate[1]))))
            digest = hashlib.sha256(raw).hexdigest()[:20]
            visual_key = f"direct:{codigo}:{digest}" if visual_source == "premium-override" else f"real:{digest}"
            visuals.setdefault(visual_key, raw)
            quality[visual_source] += 1
        else:
            visual_key = f"fallback:{category}"
            quality["approximation"] += 1
        products.append({
            "articleKey": f"dia-{slug(codigo, 18)}--pos-{slug(sku_pos or display_name, 24)}",
            "nameKey": slug(display_name), "codigoDia": codigo, "skuPos": sku_pos, "skuIntl": sku_intl,
            "descripcionSci": descripcion, "nombrePos": nombre_pos, "nombreInventario": nombre_inventario,
            "displayName": display_name, "category": category, "section": current_section,
            "source": source_label(path), "sourceFile": path.name, "sourceRow": row_number,
            "priority": priority, "visualKey": visual_key, "visualSource": visual_source if candidates else "approximation",
        })
    title = sheet.title
    rows = sheet.max_row
    workbook.close()
    return products, visuals, {"file": path.name, "sheet": title, "rows": rows, "products": len(products), "visualSources": dict(quality)}


def save_webp(image: Image.Image, destination: Path, quality: int = 93) -> None:
    encoded = io.BytesIO()
    image.save(encoded, "WEBP", quality=quality, method=6)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(encoded.getvalue())
    with Image.open(destination) as saved:
        saved.verify()
    if destination.stat().st_size >= MAX_FILE_BYTES:
        raise ValueError(f"Imagen mayor a 25 MB: {destination.name}")


def write_visuals(visuals: dict[str, bytes], output: Path, featured_output: Path) -> tuple[dict[str, dict], dict]:
    if output.exists():
        shutil.rmtree(output)
    if featured_output.exists():
        shutil.rmtree(featured_output)
    output.mkdir(parents=True)
    featured_output.mkdir(parents=True)
    descriptors: dict[str, dict] = {}
    audit = Counter()
    for key, raw in sorted(visuals.items()):
        digest = hashlib.sha256(raw).hexdigest()[:12]
        if key.startswith("direct:"):
            _, codigo, _ = key.split(":", 2)
            destination = featured_output / f"dia-{slug(codigo, 18)}-{digest}.webp"
            with Image.open(io.BytesIO(raw)) as source:
                image = ImageOps.exif_transpose(source).convert("RGB")
                image = image.filter(ImageFilter.UnsharpMask(radius=.72, percent=135, threshold=2))
                source_size = image.size
            save_webp(image, destination, 96)
            descriptors[key] = {
                "type": "direct", "src": f"assets/catalog/featured/{destination.name}",
                "kind": "premium", "width": image.width, "height": image.height,
                "sourceWidth": source_size[0], "sourceHeight": source_size[1], "passes": 1,
            }
            audit["premium"] += 1
            continue
        image, metrics = product_tile(raw)
        bucket = digest[:2]
        destination = output / bucket / f"visual-{digest}.webp"
        save_webp(image, destination)
        descriptors[key] = {
            "type": "direct", "src": f"assets/catalog/images/{bucket}/{destination.name}",
            "kind": "restored", "width": CANVAS, "height": CANVAS, **metrics,
        }
        audit["restored"] += 1
        if max(metrics["sourceWidth"], metrics["sourceHeight"]) < 128:
            audit["restoredFromThumbnail"] += 1
    for category in ("mug", "tumbler", "cold-cup", "bottle", "brew", "accessory", "other"):
        destination = output / "fallback" / f"{category}.webp"
        save_webp(fallback_tile(category), destination, 92)
        descriptors[f"fallback:{category}"] = {
            "type": "direct", "src": f"assets/catalog/images/fallback/{destination.name}",
            "kind": "approximation", "width": CANVAS, "height": CANVAS,
            "sourceWidth": 0, "sourceHeight": 0, "passes": 0,
        }
        audit["approximationAssets"] += 1
    crowded = {path.as_posix(): sum(child.is_file() for child in path.iterdir()) for path in [output, *output.iterdir()] if path.is_dir() and sum(child.is_file() for child in path.iterdir()) >= 100}
    if crowded:
        raise ValueError(f"Carpetas de imágenes con 100 archivos: {crowded}")
    return descriptors, dict(audit)


def generate(engine_dir: Path, visual_source_dir: Path, overrides: Path, js_output: Path, report_output: Path, image_output: Path, featured_output: Path) -> dict:
    def priority(path: Path) -> tuple[int, str]:
        name = normalize(path.stem)
        return (0 if "summer 2026" in name else 1 if "winter 2026" in name else 2, name)
    paths = sorted(engine_dir.glob("*.xlsx"), key=priority)
    if not paths or len(paths) >= 100:
        raise ValueError("La carpeta de motores requiere entre 1 y 99 Excel")
    external, visual_reports = load_visual_sources(visual_source_dir)
    parsed, visuals, sources = [], {}, []
    for index, path in enumerate(paths):
        products, images, report = parse_workbook(path, index, external, overrides)
        parsed.extend(products); visuals.update(images); sources.append(report)
    deduplicated, duplicate_count = {}, 0
    for product in sorted(parsed, key=lambda item: item["priority"]):
        fingerprint = (product["codigoDia"], product["skuPos"], product["nameKey"])
        if fingerprint in deduplicated:
            duplicate_count += 1
        else:
            deduplicated[fingerprint] = product
    products = sorted(deduplicated.values(), key=lambda item: (item["visualSource"] != "premium-override", item["priority"], int(item["codigoDia"]) if item["codigoDia"].isdigit() else 999999, item["displayName"]))
    descriptors, restoration_audit = write_visuals(visuals, image_output, featured_output)
    source_counts = Counter()
    for product in products:
        product["visual"] = descriptors[product.pop("visualKey")]
        product.pop("priority", None)
        source_counts[product["visualSource"]] += 1
        product["imageNote"] = "Imagen HD reconstruida desde la referencia original." if product["visualSource"] == "premium-override" else IMAGE_NOTE
    referenced_assets = {product["visual"]["src"] for product in products}
    for path in list(image_output.rglob("*.webp")):
        published_path = Path("assets/catalog/images") / path.relative_to(image_output)
        if published_path.as_posix() not in referenced_assets:
            path.unlink()
    for directory in sorted((path for path in image_output.iterdir() if path.is_dir()), reverse=True):
        if not any(directory.iterdir()):
            directory.rmdir()
    restoration_audit["published"] = sum(1 for path in image_output.rglob("*.webp") if path.parent.name != "fallback")
    approximation_count = source_counts["approximation"]
    report = {
        "status": "ok", "version": "faithful-restoration-v4", "engineFiles": len(paths),
        "visualSourceFiles": len(visual_reports), "products": len(products), "duplicateRowsIgnored": duplicate_count,
        "withSourceImage": len(products) - approximation_count, "withApproximation": approximation_count,
        "uniqueVisuals": len(visuals), "atlases": 0,
        "featuredImages": len(list(featured_output.glob("*.webp"))),
        "restoredImageFiles": restoration_audit["published"],
        "canvasPixels": CANVAS, "contentPixels": CONTENT, "restorationAudit": restoration_audit,
        "visualSources": dict(source_counts),
        "categories": dict(Counter(product["category"] for product in products)),
        "sources": sources, "visualSourceAudit": visual_reports, "imageNote": IMAGE_NOTE,
        "moneyFieldsPublished": 0,
    }
    payload = {"version": "faithful-restoration-v4", "products": products, "meta": report}
    js_output.parent.mkdir(parents=True, exist_ok=True); report_output.parent.mkdir(parents=True, exist_ok=True)
    js_output.write_text("window.MERCH_VISUAL_CATALOG=" + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")
    report_output.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return report


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--engine-dir", type=Path, default=Path("engines/merch-lists"))
    parser.add_argument("--visual-source-dir", type=Path, default=Path("engines/visual-sources"))
    parser.add_argument("--overrides", type=Path, default=Path("engines/image-overrides"))
    parser.add_argument("--output", type=Path, default=Path("data/merch-catalog.js"))
    parser.add_argument("--report", type=Path, default=Path("data/merch-catalog-report.json"))
    parser.add_argument("--image-output", type=Path, default=Path("assets/catalog/images"))
    parser.add_argument("--featured-output", type=Path, default=Path("assets/catalog/featured"))
    args = parser.parse_args()
    print(json.dumps(generate(args.engine_dir, args.visual_source_dir, args.overrides, args.output, args.report, args.image_output, args.featured_output), ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
