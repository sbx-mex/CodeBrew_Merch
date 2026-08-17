#!/usr/bin/env python3
"""Auditoría integral y reproducible de CodeBrew PWA."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from html.parser import HTMLParser
from pathlib import Path

from PIL import Image
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PERFORMANCE_BUDGETS = {
    "index.html": 35_000,
    "styles.css": 45_000,
    "catalog.css": 30_000,
    "app.js": 127_000,
    "data/products.js": 500_000,
    "data/woe.js": 1_200_000,
    "data/merch-catalog.js": 1_000_000,
    "data/stock-config.js": 15_000,
    "data/ui-config.js": 10_000,
}
REQUIRED_SHEETS = {"Base_Campaña", "Discovery", "Homologados", "Essentials"}
OBSOLETE_ALLOWLIST = (
    Path("products.js"),
    Path("icon-512.png"),
    Path("VALIDACION_CORRECCION.md"),
    Path("ELIMINAR_OBSOLETOS.txt"),
)
GENERATED_TARGETS = {"data/app-audit.js", "data/app-audit.json", "data/stock-config.js", "data/ui-config.js", "data/merch-catalog.js", "data/merch-catalog-report.json"}


class HtmlAuditParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.ids: list[str] = []
        self.references: list[str] = []

    def handle_starttag(self, tag: str, attrs) -> None:
        attributes = dict(attrs)
        if attributes.get("id"):
            self.ids.append(attributes["id"])
        for name in ("src", "href"):
            value = attributes.get(name)
            if value:
                self.references.append(value)


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def check(name: str, ok: bool, detail: str, *, warning: bool = False) -> dict:
    return {"name": name, "status": "warning" if warning and ok else ("ok" if ok else "error"), "detail": detail}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def image_is_readable(path: Path) -> bool:
    try:
        with Image.open(path) as image:
            image.verify()
        return True
    except (OSError, ValueError):
        return False


def expected_photo_upload_name(value: object) -> str:
    key = re.sub(r"[^a-z0-9]", "", str(value or "").casefold())
    if not key or key in {"na", "none"}:
        return ""
    if key.isdigit():
        key = key.lstrip("0") or "0"
    return f"{key}.jpg"


def audit(root: Path) -> dict:
    report = load_json(root / "data/import-report.json")
    catalog_report = load_json(root / "data/merch-catalog-report.json")
    manifest = load_json(root / "manifest.webmanifest")
    html = (root / "index.html").read_text(encoding="utf-8")
    parser = HtmlAuditParser()
    parser.feed(html)
    workflow_update = (root / ".github/workflows/update-price-list.yml").read_text(encoding="utf-8")
    workflow_cleanup = (root / ".github/workflows/cleanup-obsolete.yml").read_text(encoding="utf-8")
    workflow_image_audit = (root / ".github/workflows/cleanup-catalog-images.yml").read_text(encoding="utf-8")
    sw_text = (root / "sw.js").read_text(encoding="utf-8")

    checks: list[dict] = []
    checks.append(check(
        "Motor Excel",
        report.get("status") == "ok" and report.get("totalValidProducts", 0) > 0,
        f"{report.get('totalValidProducts', 0):,} artículos MERCH válidos",
    ))
    present_sheets = set(report.get("sheets", {}) if isinstance(report.get("sheets"), dict) else (item.get("sheet") for item in report.get("sheets", [])))
    checks.append(check(
        "Estructura MERCH",
        REQUIRED_SHEETS.issubset(present_sheets),
        "4 pestañas operativas localizadas" if REQUIRED_SHEETS.issubset(present_sheets) else f"Faltan: {', '.join(sorted(REQUIRED_SHEETS - present_sheets))}",
    ))
    published_image_suffixes = {".jpg", ".jpeg", ".png", ".webp"}
    restored_files = sorted(
        path for path in (root / "assets/catalog/images").rglob("*")
        if path.is_file() and path.suffix.lower() in published_image_suffixes
    )
    restored_keys = [re.sub(r"_\d+$", "", path.stem).casefold() for path in restored_files]
    restored_hashes = [sha256(path) for path in restored_files]
    unique_restored_files = len(restored_keys) == len(set(restored_keys)) and len(restored_hashes) == len(set(restored_hashes))
    featured_files = sorted((root / "assets/catalog/featured").glob("*.webp"))
    engines = sorted((root / "engines/merch-lists").glob("*.xlsx"))
    visual_sources = sorted((root / "engines/visual-sources").glob("*.zip"))
    image_overrides = sorted((root / "engines/image-overrides").rglob("*.*"))
    catalog_text = (root / "data/merch-catalog.js").read_text(encoding="utf-8")
    catalog_payload = json.loads(catalog_text.split("=", 1)[1].strip().rstrip(";"))
    catalog_products = catalog_payload.get("products", [])
    referenced_visuals = {
        visual.get("src", "")
        for product in catalog_products
        if (visual := (product.get("visual") or {})).get("src")
    }
    published_visuals = {
        path.relative_to(root).as_posix()
        for path in [*restored_files, *featured_files]
    }
    image_mode = catalog_report.get("imageMode") or catalog_payload.get("meta", {}).get("imageMode")
    clean_reset = image_mode == "clean-reset"
    common_catalog_ok = (
        catalog_report.get("status") == "ok"
        and catalog_report.get("engineFiles") == len(engines) == 3
        and catalog_report.get("visualSourceFiles") == len(visual_sources) == 3
        and catalog_report.get("products", 0) > 0
        and catalog_report.get("moneyFieldsPublished") == 0
        and '"prices"' not in catalog_text
        and catalog_report.get("atlases") == 0
        and all(path.stat().st_size < 25_000_000 for path in [*engines, *visual_sources, *image_overrides, *restored_files, *featured_files])
    )
    if image_mode == "manual-upload":
        coverage = load_json(root / "data/photo-coverage.json")
        active_count = sum(product.get("stockPriority") == "active" for product in catalog_products)
        secondary_count = sum(product.get("stockPriority") == "secondary" for product in catalog_products)
        products_with_photo = [product for product in catalog_products if product.get("visual")]
        woe_merch_products = [product for product in catalog_products if product.get("source") == "Cruce SAP + Micros"]
        active_with_photo = sum(product.get("stockPriority") == "active" for product in products_with_photo)
        lot_dirs = sorted(path for path in (root / "assets/catalog/images").iterdir() if path.is_dir())
        catalog_ok = (
            common_catalog_ok
            and [lot.name for lot in lot_dirs] == [f"lote-{number:02d}" for number in range(1, 5)]
            and all(sum(child.is_file() and child.suffix.lower() in published_image_suffixes for child in lot.iterdir()) <= 100 for lot in lot_dirs)
            and all(image_is_readable(path) for path in restored_files)
            and unique_restored_files
            and referenced_visuals.issubset(published_visuals)
            and len(published_visuals) == catalog_report.get("publishedImageFiles") == coverage.get("totals", {}).get("publishedImageFiles")
            and catalog_report.get("matchedImageFiles") == coverage.get("totals", {}).get("matchedImageFiles")
            and len(products_with_photo) == catalog_report.get("withSourceImage") == coverage.get("totals", {}).get("matchedProducts")
            and coverage.get("totals", {}).get("unmatchedImageFiles") == 0
            and coverage.get("duplicateProtection") == "one-photo-per-article"
            and coverage.get("duplicatePolicy") == "keep-first-lot-ignore-later"
            and coverage.get("version") == "manual-upload-v8-stock-ranked"
            and coverage.get("postPublishAudit") == {"status": "ok", "folders": 4, "images": len(restored_files), "duplicates": 0}
            and all(product.get("photoUploadName") == expected_photo_upload_name(product.get("codigoDia")) for product in catalog_products)
            and all(product.get("visualSource") == "manual-upload" for product in products_with_photo)
            and all(product.get("visualSource") == "pending-upload" for product in catalog_products if not product.get("visual"))
            and active_count == catalog_report.get("activeStockProducts") == catalog_payload.get("meta", {}).get("activeStockProducts")
            and active_with_photo == catalog_report.get("activeWithPhoto") == catalog_payload.get("meta", {}).get("activeWithPhoto")
            and secondary_count == catalog_report.get("secondaryProducts") == catalog_payload.get("meta", {}).get("secondaryProducts")
        )
        catalog_detail = f"{len(catalog_products):,} artículos; {len(restored_files)} fotos en 4 lotes; todos los archivos relacionados y {active_with_photo} artículos activos con foto"
    elif clean_reset:
        active_count = sum(product.get("stockPriority") == "active" for product in catalog_products)
        secondary_count = sum(product.get("stockPriority") == "secondary" for product in catalog_products)
        catalog_ok = (
            common_catalog_ok
            and not referenced_visuals
            and not published_visuals
            and all(product.get("visual") is None for product in catalog_products)
            and all(product.get("visualSource") == "pending-upload" for product in catalog_products)
            and active_count > 0
            and active_count == catalog_report.get("activeStockProducts") == catalog_payload.get("meta", {}).get("activeStockProducts")
            and secondary_count == catalog_report.get("secondaryProducts") == catalog_payload.get("meta", {}).get("secondaryProducts")
        )
        catalog_detail = f"{len(catalog_products):,} artículos; catálogo visual limpio; {active_count:,} con prioridad de inventario"
    else:
        catalog_ok = (
            common_catalog_ok
            and catalog_report.get("version") == "faithful-restoration-v4"
            and catalog_report.get("withSourceImage", 0) > 800
            and catalog_report.get("canvasPixels", 0) >= 768
            and catalog_report.get("restoredImageFiles") >= 800
            and catalog_report.get("featuredImages") == len(featured_files) >= 1
            and all(image_is_readable(path) for path in [*restored_files, *featured_files])
            and referenced_visuals == published_visuals
        )
        catalog_detail = f"{catalog_report.get('products', 0):,} artículos, {catalog_report.get('restoredImageFiles', 0):,} fotografías restauradas individualmente y {len(featured_files)} imagen HD"
    checks.append(check("Catálogo visual", catalog_ok, catalog_detail))
    control_csv = root / "data/Listado_Codigo_Dia_Fotos.csv"
    control_xlsx = root / "Control_Fotos_CodeBrew.xlsx"
    codes = {str(product.get("codigoDia")) for product in catalog_products if product.get("codigoDia")}
    codes_with_photo = {str(product.get("codigoDia")) for product in catalog_products if product.get("codigoDia") and product.get("visual")}
    csv_rows: list[dict] = []
    if control_csv.is_file():
        with control_csv.open("r", encoding="utf-8-sig", newline="") as stream:
            csv_rows = list(csv.DictReader(stream))
    workbook_sheets: list[str] = []
    if control_xlsx.is_file():
        workbook = load_workbook(control_xlsx, read_only=True, data_only=False)
        workbook_sheets = workbook.sheetnames
        workbook.close()
    photo_flags = [row.get("ESTADO FOTO") == "CON FOTO" for row in csv_rows]
    photographed_stock = [
        float(row.get("EXISTENCIA") or 0)
        for row in csv_rows
        if row.get("ESTADO FOTO") == "CON FOTO" and row.get("PRIORIDAD") == "ACTIVO"
    ]
    control_ok = (
        len(csv_rows) == len(codes)
        and workbook_sheets == ["Control de fotos"]
        and all(row.get("CÓDIGO DÍA") for row in csv_rows)
        and sum(row.get("ESTADO FOTO") == "CON FOTO" for row in csv_rows) == len(codes_with_photo)
        and photo_flags == sorted(photo_flags, reverse=True)
        and photographed_stock == sorted(photographed_stock, reverse=True)
    )
    checks.append(check(
        "Listado y Excel de fotos",
        control_ok,
        f"{len(csv_rows)} códigos Día; {sum(row.get('ESTADO FOTO') == 'CON FOTO' for row in csv_rows)} con foto; una pestaña de control",
    ))
    if image_mode == "manual-upload":
        coverage = load_json(root / "data/photo-coverage.json")
        coverage_ok = (
            coverage.get("status") == "ok"
            and coverage.get("version") == "manual-upload-v8-stock-ranked"
            and coverage.get("postPublishAudit", {}).get("status") == "ok"
            and coverage.get("postPublishAudit", {}).get("folders") == 4
            and coverage.get("postPublishAudit", {}).get("duplicates") == 0
            and coverage.get("totals", {}).get("unmatchedImageFiles") == 0
            and all(row.get("status") in {"matched", "ignored-duplicate-article"} for row in coverage.get("files", []))
            and coverage.get("matchOrder") == ["codigoDia", "skuIntl"]
            and coverage.get("packing") == "fill-lote-01-first-then-02-03-04"
            and coverage.get("crossCheck") == ["SAP", "Catalogo Micros", "Base_Campaña", "Discovery", "Homologados", "Essentials"]
            and len(woe_merch_products) == catalog_report.get("appendedWoeProducts") == coverage.get("totals", {}).get("appendedWoeProducts")
        )
        checks.append(check(
            "Control de fotos",
            coverage_ok,
            f"{coverage.get('totals', {}).get('matchedImageFiles', 0)} fotos relacionadas; cruce exacto por Código Día o SKU Intl",
        ))
    else:
        cross_checked = sum(source.get("visualSources", {}).get("crossChecked", 0) for source in catalog_report.get("sources", []))
        premium = catalog_report.get("visualSources", {}).get("premium-override", 0)
        checks.append(check(
            "Doble auditoría visual",
            cross_checked >= 500 and premium >= 1 and catalog_report.get("visualSourceAudit"),
            f"{cross_checked:,} imágenes cotejadas Excel/HTML; {premium} reconstrucción premium verificada",
        ))
    woe = report.get("woe", {})
    woe_ok = woe.get("catalogRows", 0) > 0 and woe.get("microsRows", 0) > 0 and woe.get("microsCountGroups", 0) > 0 and woe.get("exactSapDuplicatesIgnored", 0) == 0
    checks.append(check(
        "Integridad WOE",
        woe_ok,
        f"{woe.get('catalogRows', 0):,} relaciones; {woe.get('operationalSapMicrosMatch', 0):,} cruces SAP + Micros",
    ))
    pdf_config_text = (root / "data/woe-pdf-config.js").read_text(encoding="utf-8")
    pdf_config = json.loads(pdf_config_text.split("=", 1)[1].strip().rstrip(";"))
    stock_config_text = (root / "data/stock-config.js").read_text(encoding="utf-8")
    stock_config = json.loads(stock_config_text.split("=", 1)[1].strip().rstrip(";"))
    ui_config_text = (root / "data/ui-config.js").read_text(encoding="utf-8")
    ui_config = json.loads(ui_config_text.split("=", 1)[1].strip().rstrip(";"))
    export_keys = [column.get("key") for column in pdf_config.get("columns", [])]
    stock_export_keys = [column.get("key") for column in stock_config.get("columns", [])]
    stock_security_ok = all(token in (root / "app.js").read_text(encoding="utf-8") for token in (
        "validateStockReading", "stockConfirmed", "signature!=='%PDF-'", "rememberConfirmedStock", "await generateStockPdf()", "detectStockLayout", "stockTokenIndex", "stockLoadToken", "stockMatchCache", "yieldToMain", "setStockBusy",
        "detectInventoryDocument", "parseSapInventoryPage", "matchSapInventoryRow", "generateSapInventoryPdf", "Sin valor reportado", "window.print()", "label:'CONTEO'", "'Conteo'",
    )) and (root / "assets/stock_pdf_woe.jpeg").exists()
    export_ok = (
        pdf_config.get("audit", {}).get("fit") is True
        and pdf_config.get("page", {}).get("format") == "letter"
        and pdf_config.get("page", {}).get("orientation") == "portrait"
        and pdf_config.get("version") == "letter-portrait-v3-quantity"
        and export_keys == ["descripcionSap", "nombreMicros", "codigoDia", "idWoe", "qty"]
        and stock_config.get("audit", {}).get("fit") is True
        and stock_config.get("page", {}).get("format") == "letter"
        and stock_config.get("page", {}).get("orientation") == "portrait"
        and stock_config.get("version") == "stock-on-hand-v4-adaptive"
        and stock_config.get("parser", {}).get("adaptiveLayout") is True
        and stock_export_keys == ["codigoDia", "idWoe", "descripcionSap", "nombreMicros", "unidad", "qty", "estado"]
        and float(stock_config.get("parser", {}).get("zeroTolerance", 0)) >= 0.049
        and stock_security_ok
        and ui_config.get("version") == "operational-flow-v1"
        and len(ui_config.get("flows", {}).get("woe", [])) == 3
        and len(ui_config.get("flows", {}).get("stock", [])) == 3
    )
    checks.append(check(
        "Exportación PDF",
        export_ok,
        "WOE + Stock On Hand + HTML/PDF SAP; lectura separada, cruce e impresión segura",
    ))
    duplicate_ids = sorted({value for value in parser.ids if parser.ids.count(value) > 1})
    required_ids = {"mainContent", "modeMenu", "modeBack", "connectionStatus", "consulta", "woe", "etiquetado", "woeFlow", "woeNextAction", "woeSearch", "microsCatalogResults", "catalogFilters", "catalogSummary", "catalogGrid", "catalogLoadMore", "woeResults", "stockPanel", "stockFlow", "stockNextAction", "stockStoreInput", "stockAttach", "stockPdfInput", "stockIncludeZero", "stockProgress", "stockExport", "stockExcel", "stockPrint", "stockResults", "stockConfirmDialog", "stockConfirmAccept", "stockConfirmExcel"}
    redundant_controls = {"woeRun", "woeCopyList", "stockUploadGuideDialog", "stockUploadGuideAccept"}.intersection(parser.ids)
    app_text = (root / "app.js").read_text(encoding="utf-8")
    operational_tokens = ("ensureQrious", "persistWoeSelection", "restoreWoeSelection", "updateWoeFlow", "updateStockFlow", "renderCatalog", "scheduleCatalogRender", "requestAnimationFrame", "selectAppMode", "showHome", "missingPhotoWhatsappUrl", "photoUploadName", "525521107475", "https://wa.me/", "whatsapp://send?phone=", "data-photo-whatsapp", "window.location.assign", "Código Día:", "Nombre sugerido del archivo:", "Toma una foto completa y legible del termo", "visual?.src))*100000000", "catalog-missing-visual", "catalogVisibleLimit = 5", "updateViaCache:'none'", "controllerchange", "registration.update()", "quantity", "Añadir al conteo", "parseSapHtml", "sourceFamily", "selectedSapSourceRows", "exportRows", "exportStockExcel", "35*1024*1024")
    redundant_catalog_tokens = ("catalog-card-top", "catalog-source", "catalog-match", "Foto disponible", "visualQualityLabel")
    checks.append(check(
        "Navegación e interfaz",
        not duplicate_ids and not redundant_controls and required_ids.issubset(parser.ids) and all(token in app_text for token in operational_tokens) and all(token not in app_text for token in redundant_catalog_tokens) and "qrious.min.js" not in html and "openCatalogVisual" not in app_text and "catalogVisualDialog" not in html and "woeSearchClear" not in html + app_text and "localStorage.clear" not in app_text and "document.cookie" not in app_text,
        f"{len(parser.ids)} controles con ID único, navegación por teclado y progreso accesible"
        if not duplicate_ids and not redundant_controls
        else f"Revisar controles: {', '.join(sorted(set(duplicate_ids) | redundant_controls))}",
    ))
    local_refs = [value.split("?", 1)[0] for value in parser.references if not re.match(r"^(?:https?:|mailto:|tel:|#)", value)]
    missing_refs = sorted(value for value in set(local_refs) if value and value.lstrip("./") not in GENERATED_TARGETS and not (root / value.lstrip("./")).exists())
    checks.append(check(
        "Referencias locales",
        not missing_refs,
        "Todos los recursos HTML existen" if not missing_refs else f"Faltan: {', '.join(missing_refs)}",
    ))
    icon_paths = [root / str(icon.get("src", "")).lstrip("./") for icon in manifest.get("icons", [])]
    pwa_ok = manifest.get("display") == "standalone" and manifest.get("start_url") == "./" and icon_paths and all(path.exists() for path in icon_paths)
    checks.append(check(
        "Manifest PWA",
        bool(pwa_ok),
        f"{len(icon_paths)} iconos y modo instalable",
    ))
    shell_block = sw_text.split("const APP_SHELL = [", 1)[1].split("];", 1)[0] if "const APP_SHELL = [" in sw_text else ""
    shell_refs = re.findall(r"['\"](\./[^'\"]+)['\"]", shell_block)
    missing_shell = [value for value in shell_refs if value != "./" and value[2:] not in GENERATED_TARGETS and not (root / value[2:]).exists()]
    sw_ok = bool(shell_refs) and not missing_shell and "Lista_Precios_Base.xlsx" not in shell_block and "navigationPreload" in sw_text and "SKIP_WAITING" in sw_text and "key.startsWith('codebrew-')" in sw_text and "isCoreResource" in sw_text and "freshFirst" in sw_text and "cache:'no-cache'" in sw_text
    checks.append(check(
        "Caché y modo offline",
        sw_ok,
        f"{len(shell_refs)} recursos esenciales; Excel excluido del arranque" if sw_ok else f"Recursos faltantes: {', '.join(missing_shell)}",
    ))
    workflow_ok = (
        all(token in workflow_update + workflow_cleanup for token in ("actions/checkout@v5", "actions/setup-python@v6"))
        and "engines/**" in workflow_update
        and "scripts/build_all.py" in workflow_update
        and "unittest discover" in workflow_update
        and "git add --all assets/catalog" in workflow_update
        and "cleanup_catalog_images.py --apply" not in workflow_update
        and "workflow_run:" in workflow_cleanup
        and "github.event.workflow_run.conclusion == 'success'" in workflow_cleanup
        and "pip install --requirement scripts/requirements.txt" in workflow_cleanup
        and "scripts/cleanup_obsolete.py --apply" in workflow_cleanup
        and "contents: read" in workflow_image_audit
        and "--apply" not in workflow_image_audit
        and "git push" not in workflow_image_audit
    )
    cleanup_candidates = [path.as_posix() for path in OBSOLETE_ALLOWLIST if (root / path).exists()]
    checks.append(check(
        "Workflows y obsoletos",
        workflow_ok,
        f"Actualización y limpieza protegidas; {len(cleanup_candidates)} candidatos controlados",
        warning=bool(cleanup_candidates),
    ))
    project_files = [path for path in root.rglob("*") if path.is_file() and not any(part in {".git", ".codebrew-build", "__pycache__"} for part in path.relative_to(root).parts)]
    oversized = [path.relative_to(root).as_posix() for path in project_files if path.stat().st_size >= 25_000_000]
    crowded = []
    for directory in [root, *(path for path in root.rglob("*") if path.is_dir())]:
        if any(part in {".git", ".codebrew-build", "__pycache__"} for part in directory.relative_to(root).parts):
            continue
        try:
            count = sum(child.is_file() for child in directory.iterdir())
        except FileNotFoundError:
            # Una carpeta temporal de publicación puede desaparecer durante
            # el recorrido; no forma parte del artefacto final.
            continue
        if count > 100:
            crowded.append(f"{directory.relative_to(root).as_posix() or '.'} ({count})")
    checks.append(check(
        "Límites GitHub",
        not oversized and not crowded,
        "Ningún archivo alcanza 25 MB y ninguna carpeta contiene 100 archivos" if not oversized and not crowded else f"Grandes: {oversized}; carpetas: {crowded}",
    ))
    sizes = {path: (root / path).stat().st_size for path in PERFORMANCE_BUDGETS}
    over_budget = {path: size for path, size in sizes.items() if size > PERFORMANCE_BUDGETS[path]}
    checks.append(check(
        "Presupuesto de rendimiento",
        not over_budget,
        f"{sum(sizes.values()) / 1024:.0f} KB auditados en archivos críticos" if not over_budget else "Fuera de presupuesto: " + ", ".join(over_budget),
    ))

    errors = sum(item["status"] == "error" for item in checks)
    warnings = sum(item["status"] == "warning" for item in checks)
    return {
        "status": "ok" if not errors else "error",
        "auditedAtUtc": report.get("sourceModifiedAtUtc") or report.get("generatedAtUtc") or "unknown",
        "sourceFile": report.get("sourceFile", "Lista_Precios_Base.xlsx"),
        "sourceSha256": sha256(root / "Lista_Precios_Base.xlsx"),
        "checksTotal": len(checks),
        "checksOk": sum(item["status"] == "ok" for item in checks),
        "warnings": warnings,
        "errors": errors,
        "checks": checks,
        "performance": {"sizes": sizes, "budgets": PERFORMANCE_BUDGETS},
        "cleanupCandidates": cleanup_candidates,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=ROOT)
    parser.add_argument("--output", type=Path, default=Path("data/app-audit.json"))
    parser.add_argument("--js-output", type=Path, default=Path("data/app-audit.js"))
    args = parser.parse_args()
    root = args.root.resolve()
    result = audit(root)
    output = args.output if args.output.is_absolute() else root / args.output
    js_output = args.js_output if args.js_output.is_absolute() else root / args.js_output
    output.parent.mkdir(parents=True, exist_ok=True)
    js_output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    js_output.write_text("window.APP_AUDIT = " + json.dumps(result, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")
    for item in result["checks"]:
        if item["status"] in {"error", "warning"}:
            level = "error" if item["status"] == "error" else "warning"
            print(f"::{level} title={item['name']}::{item['detail']}")
    print(json.dumps({key: result[key] for key in ("status", "checksTotal", "checksOk", "warnings", "errors")}, ensure_ascii=False))
    return 0 if result["status"] == "ok" else 1


if __name__ == "__main__":
    raise SystemExit(main())
