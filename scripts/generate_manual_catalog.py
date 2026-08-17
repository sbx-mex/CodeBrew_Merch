#!/usr/bin/env python3
"""Genera rápidamente el catálogo de artículos; las fotos se integran después."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


FIELD_ALIASES = {
    "skuIntl": {"sku intl"}, "codigoDia": {"codigo dia"},
    "descripcion": {"descripcion sci", "descripcion"}, "nombrePos": {"nombre pos"},
    "nombreInventario": {"nombre inventario"}, "skuPos": {"sku pos"}, "imagen": {"imagen"},
}
REQUIRED_FIELDS = {"codigoDia", "descripcion", "nombrePos", "nombreInventario", "skuPos", "imagen"}


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", text.casefold()).strip()


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def identifier(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean(value)


def slug(value: object, limit: int = 52) -> str:
    return normalize(value).replace(" ", "-")[:limit].strip("-") or "articulo"


def source_label(path: Path) -> str:
    name = normalize(path.stem)
    if "summer 2026" in name:
        return "Summer 2026"
    if "winter 2026" in name or "core winter" in name:
        return "Core Winter 2026"
    if "generico" in name:
        return "Genéricos homologados"
    return clean(path.stem)


def category_for(value: object) -> str:
    key = normalize(value)
    if any(token in key for token in ("cold cup", "cld cup", "ccup")):
        return "cold-cup"
    if any(token in key for token in ("water bottle", "wtr btl", "bottle", "btl")):
        return "bottle"
    if any(token in key for token in ("tumbler", "tmblr", "tumb ")):
        return "tumbler"
    if any(token in key for token in ("mug", "taza", "ceramic", "cermc")):
        return "mug"
    if any(token in key for token in ("bag", "tote", "key ring", "iman", "magnet")):
        return "accessory"
    return "other"


def locate_headers(sheet) -> tuple[int, dict[str, int]]:
    for row_number, values in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), values_only=True), start=1):
        headers = [normalize(value) for value in values]
        if "codigo dia" not in headers:
            continue
        positions: dict[str, int] = {}
        for field, aliases in FIELD_ALIASES.items():
            matches = [index + 1 for index, header in enumerate(headers) if header in aliases]
            if len(matches) > 1:
                raise ValueError(f"{sheet.title}: encabezado duplicado {field}")
            if matches:
                positions[field] = matches[0]
        if missing := REQUIRED_FIELDS.difference(positions):
            raise ValueError(f"{sheet.title}: faltan encabezados {sorted(missing)}")
        return row_number, positions
    raise ValueError(f"{sheet.title}: no se localizó el encabezado Código Día")


def parse_workbook(path: Path, priority: int) -> tuple[list[dict], dict]:
    workbook = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    sheet = workbook.active
    header_row, positions = locate_headers(sheet)
    products: list[dict] = []
    current_section = "Catálogo"
    for row_number, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        def value(field: str, default_column: int | None = None):
            column = positions.get(field, default_column or 1)
            return values[column - 1] if column and column <= len(values) else None

        codigo = identifier(value("codigoDia"))
        descripcion = clean(value("descripcion"))
        sku_pos = identifier(value("skuPos"))
        if not codigo or not (descripcion or sku_pos):
            nonempty = [clean(value) for value in values if clean(value)]
            if len(nonempty) == 1:
                current_section = nonempty[0]
            continue
        sku_intl = identifier(value("skuIntl", 1))
        nombre_pos = clean(value("nombrePos"))
        nombre_inventario = clean(value("nombreInventario"))
        display_name = nombre_inventario or nombre_pos or descripcion
        products.append({
            "articleKey": f"dia-{slug(codigo, 18)}--pos-{slug(sku_pos or display_name, 24)}",
            "nameKey": slug(display_name), "codigoDia": codigo, "skuPos": sku_pos, "skuIntl": sku_intl,
            "descripcionSci": descripcion, "nombrePos": nombre_pos, "nombreInventario": nombre_inventario,
            "displayName": display_name, "category": category_for(f"{descripcion} {display_name}"),
            "section": current_section, "source": source_label(path), "sourceFile": path.name,
            "sourceRow": row_number, "priority": priority, "visualSource": "pending-upload",
            "visual": None, "imageNote": "Foto pendiente de carga.",
        })
    report = {"file": path.name, "sheet": sheet.title, "rows": sheet.max_row, "products": len(products)}
    workbook.close()
    return products, report


def generate(engine_dir: Path, visual_source_dir: Path, output: Path, report_output: Path) -> dict:
    def priority(path: Path) -> tuple[int, str]:
        name = normalize(path.stem)
        return (0 if "summer 2026" in name else 1 if "winter 2026" in name else 2, name)

    paths = sorted(engine_dir.glob("*.xlsx"), key=priority)
    if not paths or len(paths) >= 100:
        raise ValueError("La carpeta de motores requiere entre 1 y 99 Excel")
    parsed: list[dict] = []
    sources: list[dict] = []
    for index, path in enumerate(paths):
        products, source_report = parse_workbook(path, index)
        parsed.extend(products)
        sources.append(source_report)
    deduplicated: dict[tuple[str, str, str], dict] = {}
    duplicates = 0
    for product in sorted(parsed, key=lambda item: item["priority"]):
        fingerprint = (product["codigoDia"], product["skuPos"], product["nameKey"])
        if fingerprint in deduplicated:
            duplicates += 1
        else:
            deduplicated[fingerprint] = product
    products = sorted(deduplicated.values(), key=lambda item: (item["priority"], int(item["codigoDia"]) if item["codigoDia"].isdigit() else 999999, item["displayName"]))
    for product in products:
        product.pop("priority", None)
    visual_sources = list(visual_source_dir.glob("*.zip"))
    report = {
        "status": "ok", "version": "manual-catalog-v1", "engineFiles": len(paths),
        "visualSourceFiles": len(visual_sources), "products": len(products),
        "duplicateRowsIgnored": duplicates, "withSourceImage": 0, "withApproximation": 0,
        "uniqueVisuals": 0, "atlases": 0, "featuredImages": 0, "restoredImageFiles": 0,
        "publishedImageFiles": 0, "moneyFieldsPublished": 0, "imageMode": "manual-upload",
        "categories": dict(Counter(product["category"] for product in products)), "sources": sources,
        "visualSourceAudit": [], "imageNote": "Una fotografía manual por artículo.",
    }
    payload = {"version": "manual-catalog-v1", "products": products, "meta": report}
    output.parent.mkdir(parents=True, exist_ok=True)
    report_output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("window.MERCH_VISUAL_CATALOG=" + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")
    report_output.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return report


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--engine-dir", type=Path, default=Path("engines/merch-lists"))
    parser.add_argument("--visual-source-dir", type=Path, default=Path("engines/visual-sources"))
    parser.add_argument("--output", type=Path, default=Path("data/merch-catalog.js"))
    parser.add_argument("--report", type=Path, default=Path("data/merch-catalog-report.json"))
    args = parser.parse_args()
    print(json.dumps(generate(args.engine_dir, args.visual_source_dir, args.output, args.report), ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
