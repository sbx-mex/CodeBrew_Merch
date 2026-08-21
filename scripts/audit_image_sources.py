#!/usr/bin/env python3
"""Audita imágenes por lote y cruza su Código Día contra todas las hojas Excel."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image


LOT_NAMES = tuple(f"lote-{number:02d}" for number in range(1, 5))


def identifier(value: object) -> str:
    text = str(value or "").strip()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits.lstrip("0") or ("0" if digits else "")


def workbook_index(root: Path) -> tuple[dict[str, list[dict]], list[dict]]:
    index: dict[str, list[dict]] = defaultdict(list)
    audited: list[dict] = []
    paths = [root / "Lista_Precios_Base.xlsx", *sorted((root / "engines/merch-lists").glob("*.xlsx"))]
    for path in paths:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_rows = []
        for sheet in workbook.worksheets:
            rows = 0
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                rows += 1
                for column_number, value in enumerate(row, start=1):
                    key = identifier(value)
                    if 4 <= len(key) <= 10:
                        index[key].append({
                            "workbook": path.relative_to(root).as_posix(),
                            "sheet": sheet.title,
                            "cell": f"R{row_number}C{column_number}",
                        })
            sheet_rows.append({"sheet": sheet.title, "rows": rows})
        workbook.close()
        audited.append({"workbook": path.relative_to(root).as_posix(), "sheets": sheet_rows})
    return index, audited


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path("."))
    parser.add_argument("--images", type=Path, default=Path("assets/catalog/images"))
    parser.add_argument("--output", type=Path, default=Path("data/image-source-audit.json"))
    args = parser.parse_args()
    root = args.root.resolve()
    image_root = args.images if args.images.is_absolute() else root / args.images
    index, workbooks = workbook_index(root)
    seen_codes: dict[str, str] = {}
    seen_hashes: dict[str, str] = {}
    rows = []
    errors = []
    lot_counts = {}
    for lot_name in LOT_NAMES:
        lot = image_root / lot_name
        lot.mkdir(parents=True, exist_ok=True)
        images = sorted(lot.glob("*.webp"))
        lot_counts[lot_name] = len(images)
        if len(images) > 100:
            errors.append(f"{lot_name} supera 100 imágenes")
        for path in images:
            code = identifier(path.stem)
            digest = hashlib.sha256(path.read_bytes()).hexdigest()
            with Image.open(path) as image:
                size = list(image.size)
                image_format = image.format
            if code in seen_codes:
                errors.append(f"Código Día duplicado: {code} ({seen_codes[code]} y {lot_name})")
            if digest in seen_hashes:
                errors.append(f"Contenido duplicado: {path.name} y {seen_hashes[digest]}")
            if size != [960, 960] or image_format != "WEBP":
                errors.append(f"Imagen no normalizada: {lot_name}/{path.name} {size} {image_format}")
            matches = index.get(code, [])
            if not matches:
                errors.append(f"Código Día sin cruce Excel: {code}")
            seen_codes[code] = lot_name
            seen_hashes[digest] = f"{lot_name}/{path.name}"
            rows.append({"codigoDia": code, "file": f"{lot_name}/{path.name}", "excelMatches": matches})
    payload = {
        "status": "ok" if not errors else "error",
        "imageFiles": len(rows),
        "uniqueCodes": len(seen_codes),
        "lotCounts": lot_counts,
        "workbooks": workbooks,
        "files": rows,
        "errors": errors,
    }
    output = args.output if args.output.is_absolute() else root / args.output
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({key: payload[key] for key in ("status", "imageFiles", "uniqueCodes", "lotCounts", "errors")}, ensure_ascii=False))
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
